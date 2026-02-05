import re
import os
import json
import time
import requests
from datetime import timedelta, date, datetime, UTC
from collections import defaultdict
from anthropic import Anthropic
from config import HEADERS, REQUEST_DELAY, ANTHROPIC_MODEL_8K, MAX_8K_HTML_BYTES
from utils import lookup_cik_from_ticker, parse_date


# === Claude API telemetry ===
# Pricing per million tokens (Sonnet 4, as of 2025)
_CLAUDE_INPUT_COST_PER_M = 3.00
_CLAUDE_OUTPUT_COST_PER_M = 15.00


def log_claude_api(
    ticker, year, quarter, model, input_tokens, output_tokens, duration_sec, status, error_msg=None
):
    """Log Claude API call metrics to usage_logs/claude_api_log.jsonl."""
    cost_usd = (
        (input_tokens / 1_000_000) * _CLAUDE_INPUT_COST_PER_M
        + (output_tokens / 1_000_000) * _CLAUDE_OUTPUT_COST_PER_M
    )
    record = {
        "timestamp": datetime.now(UTC).isoformat(),
        "ticker": ticker,
        "year": year,
        "quarter": quarter,
        "model": model,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "duration_sec": round(duration_sec, 2),
        "cost_usd": round(cost_usd, 4),
        "status": status,
    }
    if error_msg:
        record["error"] = error_msg

    os.makedirs("usage_logs", exist_ok=True)
    with open("usage_logs/claude_api_log.jsonl", "a") as f:
        f.write(json.dumps(record) + "\n")
from edgar_tools import (
    fetch_recent_10q_10k_accessions,
    label_10q_accessions,
    enrich_10k_accessions_with_fiscal_year,
)


# TODO: n_limit=8 means only the 8 most recent Item 2.02 8-Ks are fetched.
# Requests for older quarters may fail if the correct 8-K falls outside this
# window. Consider increasing or making caller-configurable if needed.
def fetch_recent_8k_accessions(cik, headers, n_limit=8):
    url = f"https://data.sec.gov/submissions/CIK{cik.zfill(10)}.json"
    r = requests.get(url, headers=headers)
    r.raise_for_status()
    filings = r.json()["filings"]["recent"]

    results = []
    fallback_701 = []
    for i, form in enumerate(filings["form"]):
        if form != "8-K":
            continue
        # SEC submissions JSON has an "items" field for 8-Ks -- comma-separated item codes
        # Item 2.02 = "Results of Operations" (standard earnings release)
        # Item 7.01 = "Regulation FD Disclosure" (some companies file earnings here instead)
        items_str = filings.get("items", [""] * len(filings["form"]))[i] or ""
        if "2.02" not in items_str and "7.01" not in items_str:
            continue
        entry = {
            "accession": filings["accessionNumber"][i],
            "report_date": filings["reportDate"][i],
            "filing_date": filings["filingDate"][i],
            "items": items_str,
        }
        if "2.02" in items_str:
            results.append(entry)
            if len(results) >= n_limit:
                break
        elif "7.01" in items_str:
            fallback_701.append(entry)
    if results:
        return results
    # Fallback: use 7.01 if no 2.02 filings were found
    return fallback_701[:n_limit]


def fetch_8k_exhibit(cik, accession, headers):
    acc_nodash = accession.replace("-", "")
    index_url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc_nodash}/index.json"
    time.sleep(REQUEST_DELAY)
    try:
        r = requests.get(index_url, headers=headers)
        r.raise_for_status()
    except requests.RequestException:
        return None, None
    directory = r.json().get("directory", {}).get("item", [])

    exhibit_name = None

    # Priority 1: type field -- prefer EX-99.1 over EX-99.2, HTML only (skip PDFs)
    ex99_by_type = []
    for item in directory:
        item_type = item.get("type", "").upper()
        name_lower = item.get("name", "").lower()
        if item_type.startswith("EX-99") and name_lower.endswith((".htm", ".html")):
            ex99_by_type.append((item_type, item["name"]))
    # Sort numerically: extract the suffix after "EX-99." so EX-99.2 < EX-99.10
    def _ex99_sort_key(pair):
        m = re.search(r"EX-99\.?(\d+)", pair[0], re.IGNORECASE)
        return int(m.group(1)) if m else 999
    ex99_by_type.sort(key=_ex99_sort_key)
    if ex99_by_type:
        exhibit_name = ex99_by_type[0][1]

    # Priority 2: filename pattern (HTML only)
    if not exhibit_name:
        for item in directory:
            name_lower = item.get("name", "").lower()
            if ("ex99" in name_lower or "ex-99" in name_lower) and name_lower.endswith((".htm", ".html")):
                exhibit_name = item["name"]
                break

    # Priority 3: largest .htm that isn't the 8-K cover or index
    # Only exclude specific known non-exhibit files -- don't exclude by "8k" substring
    # because some real exhibits may contain "8k" in their filename
    if not exhibit_name:
        skip_names = set()
        for item in directory:
            name_lower = item.get("name", "").lower()
            item_type = item.get("type", "").upper()
            # Skip the primary 8-K document and index files
            if item_type in ("8-K", "8-K/A"):
                skip_names.add(item["name"])
            if name_lower.endswith(("-index.htm", "-index.html", "index.htm", "index.html")):
                skip_names.add(item["name"])
        htm_files = [
            (item["name"], int(item.get("size", 0) or 0))
            for item in directory
            if item["name"].lower().endswith((".htm", ".html"))
            and item["name"] not in skip_names
        ]
        if not htm_files:
            return None, None
        exhibit_name = max(htm_files, key=lambda x: x[1])[0]

    exhibit_url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc_nodash}/{exhibit_name}"
    time.sleep(REQUEST_DELAY)
    try:
        r = requests.get(exhibit_url, headers=headers)
        r.raise_for_status()
    except requests.RequestException:
        return None, None
    return r.text, exhibit_url


def find_8k_for_period(cik, headers, year, quarter, metadata_only=False):
    # 1. Get Item 2.02 8-Ks
    eight_k_list = fetch_recent_8k_accessions(cik, headers)
    if not eight_k_list:
        return None, None, None, None

    # 2. Compute expected period-end using existing fiscal calendar logic
    #    Reuse the same approach as the 10-Q/10-K pipeline:
    #    - fetch_recent_10q_10k_accessions() gives us 10-Q and 10-K lists
    #    - label_10q_accessions() assigns quarters to 10-Qs using FY-end from 10-Ks
    #    - For the requested year/quarter, look up the corresponding period-end date
    accessions_10q, accessions_10k = fetch_recent_10q_10k_accessions(cik, headers)
    try:
        accessions_10q = label_10q_accessions(accessions_10q, accessions_10k)
    except ValueError:
        # No 10-Ks with valid dates (e.g., recent IPO) — can't determine fiscal calendar
        accessions_10q = []
    accessions_10k = enrich_10k_accessions_with_fiscal_year(accessions_10k)

    # Look up expected period-end from the fiscal calendar
    # Primary: exact match from labeled 10-Qs or 10-Ks
    # Fallback: derive expected period-end using the same fiscal-year-end
    # alignment logic the main pipeline uses (via labeled 10-Qs)
    expected_period_end = None

    if quarter == 4:
        # Q4 period-end = fiscal year-end (from 10-K reportDate)
        for k in accessions_10k:
            if k.get("year") == year:
                expected_period_end = parse_date(k.get("report_date"))
                break
        if not expected_period_end:
            # No 10-K for requested year -- use closest 10-K's FY-end, shift to requested year
            # (FY-end month/day is consistent year-to-year for the vast majority of companies)
            for k in accessions_10k:
                if k.get("fiscal_year_end"):
                    expected_period_end = k["fiscal_year_end"].replace(year=year)
                    break
    else:
        # Q1-Q3: find the exact labeled 10-Q for the requested period
        target_label = f"{quarter}Q{str(year)[-2:]}"
        for q in accessions_10q:
            if q.get("label") == target_label:
                expected_period_end = parse_date(q.get("report_date"))
                break

        # Compute FY-end info upfront (needed by both ref-10-Q shift and midpoint fallback)
        fy_end_dates = [k["fiscal_year_end"] for k in accessions_10k if k.get("fiscal_year_end")]
        fy_end_consistent = len(set((d.month, d.day) for d in fy_end_dates)) <= 1 if fy_end_dates else True
        fye_target = None
        for k in accessions_10k:
            if k.get("year") == year and k.get("fiscal_year_end"):
                fye_target = k["fiscal_year_end"]
                break
        if not fye_target and fy_end_dates:
            fye_target = fy_end_dates[0].replace(year=year)

        if not expected_period_end:
            # No 10-Q for this quarter yet -- use pipeline-style fiscal-year-end alignment:
            # 1) Identify the target fiscal year-end (from 10-Ks) for the requested year.
            # 2) Find a prior 10-Q for the same quarter with a matching fiscal-year-end month/day.
            # 3) Compute the day offset between that 10-Q's report_date and its FY-end,
            #    then apply the same offset to the target FY-end.
            if fye_target and fy_end_consistent:
                target_quarter = f"Q{quarter}"
                # Find the MOST RECENT reference 10-Q with the same quarter and FY-end month/day
                ref_candidates = []
                for q in accessions_10q:
                    if q.get("quarter") != target_quarter or not q.get("report_date"):
                        continue
                    q_fye = q.get("fiscal_year_end")
                    if q_fye and (q_fye.month, q_fye.day) == (fye_target.month, fye_target.day):
                        ref_candidates.append(q)
                ref_q = None
                if ref_candidates:
                    ref_q = max(
                        ref_candidates,
                        key=lambda q: parse_date(q.get("report_date")) or date.min,
                    )
                if ref_q:
                    ref_end = parse_date(ref_q.get("report_date"))
                    ref_fye = ref_q.get("fiscal_year_end")
                    if ref_end and ref_fye:
                        delta_days = (ref_fye - ref_end).days
                        expected_period_end = fye_target - timedelta(days=delta_days)

        # Last-resort fallback: use label_10q_accessions day-diff midpoints
        # (same ranges the pipeline uses to assign quarters)
        # Q1: 250-300 days before FY-end → midpoint 275
        # Q2: 160-200 → midpoint 180
        # Q3: 70-120 → midpoint 95
        # Only used when FY-end is consistent (or known for the exact year),
        # since the midpoints assume a stable fiscal calendar.
        if not expected_period_end and fye_target and fy_end_consistent:
            DAY_DIFF_MIDPOINTS = {1: 275, 2: 180, 3: 95}
            midpoint = DAY_DIFF_MIDPOINTS.get(quarter)
            if midpoint:
                expected_period_end = fye_target - timedelta(days=midpoint)

    if not expected_period_end:
        return None, None, None, None

    # 3. Find the EARLIEST Item 2.02/7.01 8-K filed AFTER the expected period-end
    #    (within a 150-day window) with lightweight period validation on the exhibit content.
    #    Sort ascending by filing_date so we pick the closest 8-K to the period-end,
    #    not the most recent one (the SEC list is newest-first).
    #    150 days accommodates late Q4 filers (10-K deadline is 60-90 days post FY-end,
    #    but the 8-K earnings release can occasionally trail that for smaller companies).
    MAX_8K_WINDOW_DAYS = 150
    candidates = []
    for entry in eight_k_list:
        filing_date = parse_date(entry.get("filing_date"))
        if filing_date and filing_date >= expected_period_end:
            gap = (filing_date - expected_period_end).days
            if gap <= MAX_8K_WINDOW_DAYS:
                candidates.append((filing_date, entry))
    candidates.sort(key=lambda x: x[0])  # earliest first

    for _, entry in candidates:
        if metadata_only:
            # Skip HTML download — return metadata only (used by get_filings)
            return entry, None, None, str(expected_period_end)
        html, exhibit_url = fetch_8k_exhibit(cik, entry["accession"], headers)
        if not html:
            continue
        # Lightweight validation: if we can parse a period-end date from the
        # exhibit, check it matches. If we can't parse one (regex doesn't cover
        # all phrasings), accept the 8-K anyway -- the filing_date window already
        # constrains the search to a reasonable range.
        exhibit_date = _extract_period_end_from_html(html, expected_period_end)
        if exhibit_date and abs((exhibit_date - expected_period_end).days) > 15:
            continue  # period mismatch -- skip, try next candidate
        period_end_str = str(exhibit_date or expected_period_end)
        return entry, html, exhibit_url, period_end_str

    return None, None, None, None


# Regex to extract period-end dates from exhibit HTML
# Handles: "Three months ended December 31, 2025", "Quarter ended Dec. 31, 2025",
#           "Fiscal quarter ended March 29, 2025", "13 weeks ended January 1, 2026",
#           "Three months ended December 31, 2025 (unaudited)", etc.
MONTH = (
    r"(?:January|February|March|April|May|June|July|August|September|October|November|December"
    r"|Jan\.?|Feb\.?|Mar\.?|Apr\.?|May\.?|Jun\.?|Jul\.?|Aug\.?|Sep\.?|Oct\.?|Nov\.?|Dec\.?)"
)
PERIOD_DATE_PATTERN = re.compile(
    r"(?:for\s+the\s+)?"
    r"(?:fiscal\s+)?"  # optional "fiscal" prefix
    r"(?:"
    r"(?:three|six|nine|twelve|thirteen|fourteen|52|53)\s+(?:months|weeks)\s+ended"
    r"|quarter\s+ended"
    r")\s+"
    r"(" + MONTH + r"\s+\d{1,2},?\s+\d{4})"
    r"(?:\s*\(unaudited\))?",  # optional trailing "(unaudited)"
    re.IGNORECASE,
)


def _extract_period_end_from_html(html, expected_period_end):
    """Extract a period-end date from the exhibit HTML. Returns closest match or None."""
    matches = PERIOD_DATE_PATTERN.findall(html)
    if not matches:
        return None
    parsed = []
    for m in matches:
        dt = parse_date(m)
        if dt:
            parsed.append(dt)
    if not parsed:
        return None
    # Prefer the date closest to the expected period-end
    return min(parsed, key=lambda d: abs((d - expected_period_end).days))


def _strip_html_attrs(html_text):
    """Strip all attributes from HTML tags to reduce token count.
    Keeps the tag structure (table, tr, td, th, span, etc.) but removes
    style, class, id, width, and other attributes that bloat SEC filings.

    NOTE: This also removes colspan/rowspan attributes, which can affect table
    alignment for filings that use merged cells. In practice, SEC earnings
    press releases rarely use complex cell spans, and Claude handles the
    simplified structure well. If extraction quality degrades for a specific
    company, consider preserving colspan/rowspan by switching to a more
    selective regex, e.g.:
        re.sub(r'(?!\\s(?:colspan|rowspan)=)\\s[a-z-]+=("[^"]*"|...)', '', tag)
    """
    return re.sub(r"<(\/?[a-zA-Z][a-zA-Z0-9]*)\b[^>]*\/?>", r"<\1>", html_text)


def extract_facts_from_8k(html_content, ticker, year, quarter, full_year_mode):
    # 1. Preprocess HTML
    html = re.sub(
        r"<(style|script|head)[^>]*>.*?</\1>", "", html_content, flags=re.DOTALL | re.IGNORECASE
    )
    # Strip all HTML attributes (style, class, etc.) -- biggest size reducer
    html = _strip_html_attrs(html)
    # Remove empty tags and collapse whitespace
    html = re.sub(r"<(td|th|span|div|p)>\s*</\1>", "", html, flags=re.IGNORECASE)
    html = re.sub(r"\s{2,}", " ", html)

    if len(html.encode("utf-8")) > MAX_8K_HTML_BYTES:
        # Extract just tables with surrounding context
        tables = re.findall(
            r"(.{0,200}<table>.*?</table>.{0,200})",
            html,
            flags=re.DOTALL | re.IGNORECASE,
        )
        if tables:
            html = "\n\n".join(tables)
        else:
            # No tables found — fallback to truncated original HTML
            html = html[:MAX_8K_HTML_BYTES].rsplit(" ", 1)[0]  # truncate at word boundary

    # 2. Build prompt
    period = f"FY {year}" if full_year_mode else f"Q{quarter} {year}"
    prompt = f"""This is {ticker}'s earnings press release for {period}.

Extract all financial line items from every table (quarterly, YTD,
full-year, and balance sheet) as year-over-year comparisons -- the
actual numbers, not the change -- with their labels.

Output a JSON array where each item has:
- "tag": a descriptive label for the line item. If the row label is generic
  (e.g., "Diluted", "Basic"), prepend the section or table header to
  disambiguate (e.g., "Earnings per share, diluted", "Shares outstanding, basic")
- "current": the most recent period's number for that table
- "prior": the prior-year comparison number for that same period
- "date_type": must be exactly one of these string values:
  - "Q" -- single quarter (e.g., "three months ended"). Balance sheet items ("as of" a date) ALWAYS use "Q" regardless of the date.
  - "YTD" -- year-to-date or cumulative (e.g., "nine months ended")
  - "FY" -- full fiscal year (e.g., "twelve months ended" or annual)

Numbers in parentheses are negative. Strip commas but don't apply any
scale factor. Use null for missing values.

Output ONLY the JSON array, no other text."""

    # 3. Call Claude API
    client = Anthropic()  # uses ANTHROPIC_API_KEY env var
    start_time = time.time()
    try:
        response = client.messages.create(
            model=ANTHROPIC_MODEL_8K,
            max_tokens=8192,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "text", "text": html},
                    ],
                }
            ],
        )
    except Exception as e:
        duration = time.time() - start_time
        log_claude_api(ticker, year, quarter, ANTHROPIC_MODEL_8K, 0, 0, duration, "error", str(e))
        raise ValueError(
            f"Anthropic API error for {ticker} Q{quarter} {year}: {type(e).__name__}: {e}"
        )

    duration = time.time() - start_time
    input_tokens = getattr(response.usage, "input_tokens", 0)
    output_tokens = getattr(response.usage, "output_tokens", 0)

    if not response.content:
        log_claude_api(ticker, year, quarter, ANTHROPIC_MODEL_8K, input_tokens, output_tokens, duration, "empty_response")
        raise ValueError(f"Anthropic API returned empty response for {ticker} Q{quarter} {year}")
    raw_text = response.content[0].text.strip()

    # 4. Parse JSON -- handle markdown fencing if Claude adds it
    if raw_text.startswith("```"):
        raw_text = re.sub(r"^```\w*\n?", "", raw_text)
        raw_text = re.sub(r"\n?```$", "", raw_text)
    try:
        raw_facts = json.loads(raw_text)
    except json.JSONDecodeError as e:
        # Try to salvage: find the first [ ... ] in the response
        bracket_match = re.search(r"\[.*\]", raw_text, re.DOTALL)
        if bracket_match:
            try:
                raw_facts = json.loads(bracket_match.group())
            except json.JSONDecodeError:
                log_claude_api(ticker, year, quarter, ANTHROPIC_MODEL_8K, input_tokens, output_tokens, duration, "invalid_json")
                raise ValueError(
                    f"Claude returned invalid JSON for {ticker} Q{quarter} {year}: {e}. "
                    f"First 200 chars: {raw_text[:200]}"
                )
        else:
            log_claude_api(ticker, year, quarter, ANTHROPIC_MODEL_8K, input_tokens, output_tokens, duration, "invalid_json")
            raise ValueError(
                f"Claude returned invalid JSON for {ticker} Q{quarter} {year}: {e}. "
                f"First 200 chars: {raw_text[:200]}"
            )

    if not isinstance(raw_facts, list):
        log_claude_api(ticker, year, quarter, ANTHROPIC_MODEL_8K, input_tokens, output_tokens, duration, "invalid_type")
        raise ValueError(
            f"Claude returned {type(raw_facts).__name__} instead of list for {ticker} Q{quarter} {year}. "
            f"First 200 chars: {raw_text[:200]}"
        )

    # 5. Post-process into EdgarFact schema and log success
    log_claude_api(ticker, year, quarter, ANTHROPIC_MODEL_8K, input_tokens, output_tokens, duration, "success")
    return _postprocess_facts(raw_facts)


def _coerce_numeric(val):
    """Ensure value is numeric. Handle strings with commas/parens from Claude output.

    NOTE: Values like "15%", "N/M", "—" will return None. The prompt instructs Claude
    to output raw numbers without % signs, so this should be rare. If margin metrics
    with % become common, consider stripping % and dividing by 100.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        s = val.strip().replace(",", "")
        # Handle common placeholder dashes
        if s in ("—", "-", "–", "N/A", "N/M", ""):
            return None
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return int(s) if "." not in s else float(s)
        except ValueError:
            return None
    return None


VALID_DATE_TYPES = {"Q", "YTD", "FY"}


def _postprocess_facts(raw_facts):
    """Map Claude's {tag, current, prior, date_type} to full EdgarFact schema."""
    facts = []
    for item in raw_facts:
        current = _coerce_numeric(item.get("current"))
        prior = _coerce_numeric(item.get("prior"))
        raw_dt = item.get("date_type")
        date_type = raw_dt.upper().strip() if isinstance(raw_dt, str) else None
        date_type = date_type if date_type in VALID_DATE_TYPES else None
        facts.append(
            {
                "tag": item.get("tag"),
                "date_type": date_type,
                "presentation_role": None,
                "current_period_value": current,
                "prior_period_value": prior,
                "visual_current_value": current,
                "visual_prior_value": prior,
                "end_current": None,
                "end_prior": None,
                "axis_segment": None,
                "axis_geo": None,
                "collision_flag": 0,  # computed below
            }
        )

    # Collision detection: same prior → multiple distinct currents
    # Skip prior values that are 0 or null (too common, would spike false collisions)
    prior_to_currents = defaultdict(set)
    for f in facts:
        p = f["prior_period_value"]
        c = f["current_period_value"]
        if p is not None and p != 0 and c is not None:
            prior_to_currents[p].add(c)

    colliding_priors = {p for p, currents in prior_to_currents.items() if len(currents) > 1}
    for f in facts:
        if f["prior_period_value"] in colliding_priors:
            f["collision_flag"] = 1

    return facts


def get_financials_from_8k(ticker, year, quarter, full_year_mode=False, use_cache=True):
    # === 1. Check cache ===
    cache_dir = "exports"
    cache_filename = (
        f"{ticker}_FY{str(year)[-2:]}_8k_financials.json"
        if full_year_mode
        else f"{ticker}_{quarter}Q{str(year)[-2:]}_8k_financials.json"
    )
    cache_path = os.path.join(cache_dir, cache_filename)

    # NOTE: app.py also checks this cache at the endpoint level for /api/financials.
    # This is intentionally redundant so CLI calls (which bypass Flask) also get caching.
    if use_cache and os.path.exists(cache_path):
        with open(cache_path, "r") as f:
            return json.load(f)

    # === 2. CIK lookup ===
    cik = lookup_cik_from_ticker(ticker)
    if not cik:
        return {"status": "error", "message": f"Could not find CIK for {ticker}"}

    # === 3. Find 8-K for period ===
    entry, html, exhibit_url, expected_period_end = find_8k_for_period(cik, HEADERS, year, quarter)
    if not entry:
        return {"status": "error", "message": f"No Item 2.02 8-K found for {ticker} Q{quarter} {year}"}

    # Detect value scale from HTML (informational only)
    scale_match = re.search(r"in\s+(thousands|millions|billions)", html, re.IGNORECASE)
    value_scale = scale_match.group(1).lower() if scale_match else "unknown"

    # === 4. Extract facts via Claude ===
    try:
        facts = extract_facts_from_8k(html, ticker, year, quarter, full_year_mode)
    except ValueError as e:
        return {"status": "error", "message": str(e)}

    result = {
        "status": "success",
        "metadata": {
            "ticker": ticker,
            "year": year,
            "quarter": quarter,
            "full_year_mode": full_year_mode,
            "total_facts": len(facts),
            "source": {
                "filing_type": "8-K",
                "period_end": expected_period_end or entry.get("filing_date"),
                "url": exhibit_url,
                "value_scale": value_scale,
                "cik": cik,
                "accession": entry["accession"],
            },
        },
        "facts": facts,
    }

    # === 5. Save to cache ===
    os.makedirs(cache_dir, exist_ok=True)
    with open(cache_path, "w") as f:
        json.dump(result, f)

    return result


def write_8k_facts_to_excel(facts, ticker, year, quarter, full_year_mode, excel_file):
    """Write 8-K extracted facts to Raw_data sheet, matching pipeline Excel format."""
    import openpyxl

    wb = openpyxl.load_workbook(excel_file, keep_vba=True)
    sheet = wb["Raw_data"]

    # Clear existing data
    for row in sheet["A2:E5000"]:
        for cell in row:
            cell.value = None

    # Write facts
    for i, fact in enumerate(facts):
        row_num = i + 2
        sheet.cell(row=row_num, column=1, value=fact.get("tag"))
        sheet.cell(
            row=row_num,
            column=2,
            value=fact.get("visual_current_value", fact.get("current_period_value")),
        )
        sheet.cell(
            row=row_num,
            column=3,
            value=fact.get("visual_prior_value", fact.get("prior_period_value")),
        )
        sheet.cell(row=row_num, column=4, value=fact.get("presentation_role", ""))
        sheet.cell(row=row_num, column=5, value=fact.get("collision_flag", 0))

    # Write metadata
    sheet["G1"] = "Ticker"
    sheet["H1"] = ticker
    sheet["G2"] = "Year"
    sheet["H2"] = year
    sheet["G3"] = "Quarter"
    sheet["H3"] = quarter
    sheet["G4"] = "Full Year Mode"
    sheet["H4"] = str(full_year_mode)

    wb.save(excel_file)
