#!/usr/bin/env python
# coding: utf-8

# In[ ]:


# === FULL FUNCTION TO RUN EDGAR EXTRACTOR ==========================================

def run_edgar_pipeline(
    ticker,
    year,
    quarter,
    full_year_mode,
    debug_mode,
    excel_file,
    sheet_name
):
    global TICKER, YEAR, QUARTER, FULL_YEAR_MODE, DEBUG_MODE, EXCEL_FILE, SHEET_NAME
    # Set variables from function inputs
    TICKER = ticker
    YEAR = year
    QUARTER = quarter
    FULL_YEAR_MODE = full_year_mode
    DEBUG_MODE = debug_mode
    EXCEL_FILE = excel_file
    SHEET_NAME = sheet_name

    # ... FULL pipeline logic from your Jupyter notebook pasted here, unchanged ...

    #!/usr/bin/env python
    # coding: utf-8
    
    # In[ ]:
    
    
    # === MODULE IMPORTS ===
    
    from utils import (
        metrics,
        log_metric, 
        extract_fiscal_year_end,
        extract_dimensions_from_context,
        AXIS_COLS,
        PREFIX_MAP,
        FINAL_COLS,
        lookup_cik_from_ticker,
        zip_match_in_order,
        audit_value_collisions,
        run_adaptive_match_keys,
        standardize_zip_output,
        parse_date
    )
    
    from enrich import (
        get_negated_label_concepts,
        get_concept_roles_from_presentation
    )
    
    from config import (
        TICKER_CIK_URL,
        HEADERS,
        N_10Q,
        N_10K,
        N_10Q_EXTRA,
        N_10K_EXTRA,
        REQUEST_DELAY,
        OUTPUT_METRICS_DIR,
        EXPORT_UPDATER_DIR
    )
    
    
    # In[ ]:
    
    
    # === IMPORTS ===
    
    import requests
    import re
    from datetime import datetime
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl import load_workbook
    import html
    import time
    import os
    import json
    from datetime import datetime
    
    
    # In[ ]:
    
    
    # === Add inputs to metrics dictionary ===
    metrics.update({
        "ticker": ticker,
        "year": year,
        "quarter": quarter,
        "full_year_mode": full_year_mode,
        "debug_mode": debug_mode,
        "start_time": datetime.now().isoformat()
    })
    
    
    # In[ ]:
    
    
    
    
    
    # In[ ]:
    
    # === EXTRACTION CONFIG ===
    
    CIK = lookup_cik_from_ticker(TICKER)
    FOUR_Q_MODE = (QUARTER == 4)  # 🆕 Build 4Q flag
    
    # Adjust number of filings to pull based on mode - # You might need more for 4Q builds (Q1–Q3 of both years)
    if FOUR_Q_MODE:
        N_10Q = N_10Q + N_10Q_EXTRA
        N_10K = N_10K + N_10K_EXTRA
    else:
        N_10Q = N_10Q + N_10Q_EXTRA
        N_10K = N_10K + N_10K_EXTRA
    
    # === Enforce quarter numbers == 
    
    if QUARTER not in [1, 2, 3, 4]:
        raise ValueError(f"❌ Invalid quarter value: {QUARTER}. Must be 1, 2, 3, or 4.")
    
    # === Enforce FULL_YEAR_MODE only when QUARTER == 4
    if QUARTER != 4:
        FULL_YEAR_MODE = False
    
    import os
    
    # === Check if Excel file exists ===
    if not os.path.exists(EXCEL_FILE):
        print(f"⚠️ Warning: Excel file '{EXCEL_FILE}' does not exist yet. It will need to be created or exported later.")
    else:
        print(f"✅ Excel file '{EXCEL_FILE}' found.")
    
    # === Check if CIK was properly loaded ===
    if not CIK:
        raise ValueError("❌ No valid CIK provided. Please set CIK or lookup from TICKER.")
    print(f"✅ Using CIK for {TICKER}: {CIK}")
    
    
    # Notes:
    # - Always run cells in sequence
    # - After export to Updater.xlsm, manually archive if needed
    # - Change CIK/YEAR/QUARTER above before rerunning
    
    
    # In[ ]:
    
    
    # === FETCH & PARSE FILINGS ===================================
    # Function to parse filings and label the data
    # === Categorize Periods for Extracted Facts from Filing(s) ===
    
    from datetime import datetime, timedelta
    from lxml import etree
    import pandas as pd
    import re
    
    
    # === Helper: Enrich a Filing ===
    def enrich_filing(filing):
    
        """
        Enriches a parsed EDGAR filing by categorizing all XBRL facts into time-based categories
        (e.g., current quarter, prior year), applying context dimension mapping, and labeling
        period types and axes. Returns the result as a fully labeled pandas DataFrame.
    
        Args:
            filing (dict): Parsed filing dictionary containing keys like 'facts', 'context_blocks',
                           'form', 'document_period_end', 'accession', and 'concept_roles'.
    
        Returns:
            pandas.DataFrame: Enriched DataFrame of facts with columns for:
                - Financial value tags, context, and matched periods
                - Time-based labels (Q, YTD, FY)
                - Axis assignments (segment, product, geo, etc.)
                - Presentation roles and categorization logic
    
        Notes:
            - Requires global access to prior 10-Ks and 10-Qs (results_10k, results_10q).
            - Automatically detects fiscal year boundaries based on historical filings.
            - Designed to support structured financial modeling and matching logic.
    
        Example:
            df = enrich_filing(filing)
        """
        
        # === Step 1: Build reference dates from the filings ===
        
        # Get current period end date
        doc_end_date = parse_date(filing["document_period_end"])
        
        form = filing.get("form")  # "10-K" or "10-Q"
        prior_start_date = None
        prior_end_date = None
    
        # Sort filings by period end, descending
        sorted_10k = sorted(
            [f for f in results_10k if parse_date(f.get("document_period_end"))],
            key=lambda f: parse_date(f["document_period_end"]),
            reverse=True
        )
        
        sorted_10q = sorted(
            [f for f in results_10q if parse_date(f.get("document_period_end"))],
            key=lambda f: parse_date(f["document_period_end"]),
            reverse=True
        )
            
        # === Block to prevent extraction from filings before 2019 (no XBRL) ===
        if doc_end_date < parse_date("2019-01-01"):
            raise ValueError(f"⚡ Filing date {doc_end_date} is before 2019. Sorry - this script only supports EDGAR filings from 2019 onward (inline XBRL not reliable before that). Please choose a filing from 2018 or later.")
        
        print("--------------------------------------------------")
        if filing.get("form") == "10-K":
            filing_label = f"FY{filing.get('year') % 100:02d}"
        else:
            filing_label = filing.get("label", "Unknown")
        
        print(f"\n🚀 Starting enrichment for {filing.get('form', 'Unknown')} [{filing_label}] | Period End: {filing.get('document_period_end', 'Unknown')} | Accession: {filing.get('accession', 'Unknown')}")
     
    
        # === Step 2: Get fiscal year start and end dates for current period and prior y/y period ===
        
        #Calculate fiscal year start with end date of prior 10K (prior fiscal year end date)
        
        # Sort 10-Ks by document_period_end descending
        sorted_10ks = sorted(results_10k, key=lambda x: parse_date(x["document_period_end"]), reverse=True)
    
        # Find prior 10-K end date (before doc_end_date)
        prior_10k_end_date = None
        for filing_prior in sorted_10ks:
            prior_end = parse_date(filing_prior["document_period_end"])
            if prior_end < doc_end_date:   
                prior_10k_end_date = prior_end  
                break
    
        # Fallback if not found
        if not prior_10k_end_date:
            print(f"⚠️ No prior 10-K found before {doc_end_date}. Using fallback prior 10-K end date with adjusted year.")
            prior_10k_end_date = doc_end_date.replace(year=doc_end_date.year - 1)
    
        # Use prior 10-K end to define fiscal year start
        prior_fiscal_year_end = prior_10k_end_date
        fiscal_year_start = prior_fiscal_year_end + timedelta(days=1)
    
        # Calculate prior fiscal year start with end date or 10-K before the prior 10-K (prior prior 10K end date)
        # To calculate the prior year start dates in filing
        
        prior_prior_10k_end_date = None
        for filing_prior2 in sorted_10ks:
            prior_end2 = parse_date(filing_prior2["document_period_end"])
            if prior_end2 and prior_end2 < prior_10k_end_date:
                prior_prior_10k_end_date = prior_end2
                break
        
        if not prior_prior_10k_end_date:
            print("⚠️ No second prior 10-K found — using fallback year subtraction.")
            prior_fiscal_year_start = fiscal_year_start.replace(year=fiscal_year_start.year - 1)
            
        else:
            prior_fiscal_year_start = prior_prior_10k_end_date + timedelta(days=1)
    
        # Set doc start date as fiscal year start if 10K filing
        if form == "10-K":
            doc_start_date = fiscal_year_start
            prior_start_date = prior_fiscal_year_start
            prior_end_date = prior_fiscal_year_end
    
        else:    
            
        # === Step 3: Get period start dates based on prior filings for 10Q's ===
            try:
                prior_filings = sorted(
                    results_10q + results_10k,
                    key=lambda x: parse_date(x.get("document_period_end")),
                    reverse=True
                )
        
            except Exception as e:
                print(f"⚠️ Warning: Failed to sort filings by document_period_end. Error: {e}")
                prior_filings = []  # fallback to empty list
                    
            doc_start_date = None
            for prior in prior_filings:
                try:
                    candidate_end = parse_date(prior["document_period_end"])
                    if candidate_end < doc_end_date:
                        doc_start_date = candidate_end + timedelta(days=1)
                        break
                except:
                    continue
                
            if not doc_start_date:
                doc_start_date = (doc_end_date - timedelta(days=90)).replace(day=1) #logic to use if no prior quarterly filings
                
        # === Step 4: Get prior y/y start and end dates from prior filings ===
        
        if form == "10-Q":
            quarter = filing.get("quarter")
            year = filing.get("year")
    
            prior_end_date = None
        
            if quarter and year:
                for q in results_10q:
                    if (
                        q.get("quarter") == quarter
                        and q.get("year") == (year - 1)
                    ):
                        q_end = parse_date(q["document_period_end"])
                        if q_end:
                            prior_end_date = parse_date(q_end)
                            break
    
            prior_start_date = None
            
            for prior in prior_filings:
                try:
                    candidate_end = parse_date(prior["document_period_end"])
                    if candidate_end < prior_end_date:
                        prior_start_date = candidate_end + timedelta(days=1)
                        break
                except:
                    continue
            
        # Fallback: if either value is still missing
        if not prior_start_date or not prior_end_date:
            print("⚠️ prior_start_date or prior_end_date missing — applying YoY fallback")
            prior_start_date = doc_start_date.replace(year=doc_start_date.year - 1)
            prior_end_date = doc_end_date.replace(year=doc_end_date.year - 1)
        
        print(f"\n🎯 Current Period: {doc_start_date} to {doc_end_date} with fiscal year start {fiscal_year_start}")
        print(f"🎯 Prior Period:   {prior_start_date} to {prior_end_date} with fiscal year start {prior_fiscal_year_start}")
        
        # Step 5: Build context period lookup
        period_lookup = {}
        context_blocks = filing["context_blocks"]
    
        # === NEW: Extract dimension info per contextref ===
        context_dim_map = { 
            ctx_id: extract_dimensions_from_context(ctx_html)
            for ctx_id, ctx_html in context_blocks.items()
        }
        
        for ctx_id, block in context_blocks.items():
            if not ctx_id:
                continue
            if "<xbrli:startdate>" in block.lower() and "<xbrli:enddate>" in block.lower():
                start = re.search(r"<xbrli:startdate>(.*?)</xbrli:startdate>", block, re.IGNORECASE)
                end = re.search(r"<xbrli:enddate>(.*?)</xbrli:enddate>", block, re.IGNORECASE)
                if start and end:
                    start = parse_date(start.group(1))
                    end = parse_date(end.group(1))
                    period_lookup[ctx_id] = ("duration", start, end)
            elif "<xbrli:instant>" in block.lower():
                instant = re.search(r"<xbrli:instant>(.*?)</xbrli:instant>", block, re.IGNORECASE)
                if instant:
                    instant = parse_date(instant.group(1))
                    period_lookup[ctx_id] = ("instant", instant)
    
        print(f"\n🧠 Mapped {len(period_lookup)} contextrefs to periods.")
    
        # Step 6: Enrich facts
        all_facts = []
        for fact in filing["facts"]:
            ctx = fact.get("contextref")
            tag = fact.get("tag")
            value = fact.get("value")
            
            if ctx not in period_lookup:
                continue
            
            period_info = period_lookup[ctx]
            enriched = {
                "tag": tag,
                "value": value,
                "contextref": ctx,
                "period_type": period_info[0],
                "matched_category": None,
                "start": None,
                "end": None,
                "date_type": None,
                "presentation_role": None
            }
    
            # Assign presentation role if concept exists in pre.xml map
            roles = filing.get("concept_roles", {}).get(tag, [])
            enriched["presentation_role"] = (
                "|".join(sorted(set(r.lower() for r in roles if isinstance(r, str))))
                if roles else None
            )
            
            dims = context_dim_map.get(ctx, [])
            
            # Initialize axis category columns
            axis_columns = [
                "axis_consolidation",
                "axis_segment",
                "axis_product",
                "axis_geo",
                "axis_legal_entity",
                "axis_unassigned" 
            ]
            for col in axis_columns:
                enriched[col] = None
            
            # Smart dimension assignment (no mapping)
            for d in dims:
                axis = (d.get("dimension") or "").lower()
                member = d.get("member")
            
                if "consolidation" in axis:
                    enriched["axis_consolidation"] = member
                elif "segment" in axis or "business" in axis:
                    enriched["axis_segment"] = member
                elif "product" in axis or "service" in axis:
                    enriched["axis_product"] = member
                elif "geo" in axis or "region" in axis or "country" in axis:
                    enriched["axis_geo"] = member
                elif "legal" in axis or "entity" in axis:
                    enriched["axis_legal_entity"] = member
    
            # === NEW: Catch-all for unclassified axes ===
            classified_keywords = ["consolidation", "segment", "business", "product", "service", "geo", "region", "country", "legal", "entity"]
            unclassified_dims = []
            
            for d in dims:
                axis = (d.get("dimension") or "").lower()
                member = d.get("member")
                if not any(k in axis for k in classified_keywords):
                    unclassified_dims.append(f"{axis}={member}")
            
            enriched["axis_unassigned"] = "|".join(unclassified_dims) if unclassified_dims else None
            
            if period_info[0] == "duration": #Categorizing flow values (revenues, etc.) as current or prior FY, YTD, or Q periods
                start, end = period_info[1], period_info[2]
                enriched["start"] = start
                enriched["end"] = end
    
                if filing.get("form") == "10-K":
                    if start == fiscal_year_start and end == doc_end_date:
                        enriched["matched_category"] = "current_full_year"
                    elif start == prior_fiscal_year_start and end == prior_fiscal_year_end:
                        enriched["matched_category"] = "prior_full_year"
    
                else: # 10-Q logic
                    if start == doc_start_date and end == doc_end_date:
                        enriched["matched_category"] = "current_q"
                    elif start == fiscal_year_start and end == doc_end_date:
                        enriched["matched_category"] = "current_ytd"
                    elif start == prior_start_date and end == prior_end_date:
                        enriched["matched_category"] = "prior_q"
                    elif start == prior_fiscal_year_start and end == prior_end_date:
                        enriched["matched_category"] = "prior_ytd"
                    
            elif period_info[0] == "instant": #Categorizing instant values (cash, etc.) as current or prior Q
                instant = period_info[1]
                enriched["end"] = instant
                if instant == doc_end_date:
                    enriched["matched_category"] = "current_q"
                elif instant == prior_end_date:
                    enriched["matched_category"] = "prior_q"
    
            # Categorize matched_category into simplified date_type
            mc = enriched["matched_category"]
            if mc in ["current_q", "prior_q"]:
                enriched["date_type"] = "Q"
            elif mc in ["current_ytd", "prior_ytd"]:
                enriched["date_type"] = "YTD"
            elif mc in ["current_full_year", "prior_full_year"]:
                enriched["date_type"] = "FY"
            else:
                enriched["date_type"] = None
            
            all_facts.append(enriched)
    
        print(f"\n✅ {len(all_facts)} facts extracted and enriched.")
        
        # Step 7: Build DataFrame
        df = pd.DataFrame(all_facts)
        print("\n🎯 Full categorization and enrichment complete!")
        print(f"✅ Completed enrichment for {filing.get('form', 'Unknown')} [{filing_label}] | Facts enriched: {len(all_facts)}")
        print("--------------------------------------------------")
        return df
    
    
    # In[ ]:
    
    
    # === FETCH & PARSE FILINGS ===================================
    # Fetches recent filings of company from JSON API and labels and puts 10K's and 10Q's in a list
    
    # === Fetch recent 10-Q and 10-K accessions from EDGAR ===
    def fetch_recent_10q_10k_accessions(cik, headers):
    
        """
        Retrieves recent 10-Q and 10-K filings for a specified company from the SEC's EDGAR JSON submissions API.
    
        This function accesses the SEC’s real-time company submissions feed, parses recent filings,
        and extracts metadata for Form 10-Q and 10-K filings. It returns two lists containing
        the accession number, report date, and form type for each matching filing.
    
        Note: This method only returns filings that are currently available in the SEC JSON feed,
        which may be limited to the most recent 250–1000 filings depending on company activity.
    
        Parameters:
            cik (str): The 10-digit Central Index Key (CIK) of the company. Leading zeros will be padded if missing.
            headers (dict): HTTP headers for the request, including a required 'User-Agent' field per SEC policy.
    
        Returns:
            tuple: Two lists:
                - accessions_10q (list of dict): Each dict contains 'accession', 'report_date', and 'form' for a 10-Q.
                - accessions_10k (list of dict): Each dict contains 'accession', 'report_date', and 'form' for a 10-K.
    
        Example call:
            fetch_recent_10q_10k_accessions("0000320193", headers)
    
        Example output:
            {'accession': '0001193125-23-123456', 'report_date': '2023-10-31', 'form': '10-Q'}
        """
        
        url = f"https://data.sec.gov/submissions/CIK{cik.zfill(10)}.json"
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        data = r.json()
        
        filings = data["filings"]["recent"]
    
        # === Validate filings structure ===
        required_keys = ["form", "accessionNumber", "reportDate"]
        if not all(k in filings for k in required_keys):
            raise ValueError("❌ SEC filings JSON missing expected fields (form, accessionNumber, reportDate).")
    
        # === Proceed safely after validation ===
        forms = filings["form"]
        accessions = filings["accessionNumber"]
        report_dates = filings["reportDate"]
    
        accessions_10q = []
        accessions_10k = []
    
        for i, form in enumerate(forms):
            entry = {
                "accession": accessions[i],
                "report_date": report_dates[i],
                "form": form
            }
    
            if form == "10-Q":
                accessions_10q.append(entry)
            elif form == "10-K":
                accessions_10k.append(entry)
    
        print(f"✅ Found {len(accessions_10q)} 10-Q accessions (filing submissions)")
        print(f"✅ Found {len(accessions_10k)} 10-K accessions (filing submissions)")
        
        return accessions_10q, accessions_10k
    
    def filter_filings_by_year(accessions, max_year, n_limit):
        """
        Filters filings to only include those with report_date ≤ max_year and returns up to n_limit entries.
        
        Parameters:
            accessions (list of dict): Filings with keys including 'report_date'.
            max_year (int): Latest year to include.
            n_limit (int): Max number of entries to return.
        
        Returns:
            list of dict: Filtered filings.
        """
        filtered = []
        for entry in accessions:
            date_str = entry.get("report_date", "")
            if not date_str or date_str.strip() == "":
                continue
            try:
                yr = int(date_str.split("-")[0])
            except:
                continue
            if yr > max_year:
                continue
            filtered.append(entry)
            if len(filtered) >= n_limit:
                break
        return filtered
    
    # === FETCH 10Q/10K ACCESSIONS ===
    accessions_10q, accessions_10k = fetch_recent_10q_10k_accessions(CIK, HEADERS)
    print(accessions_10q[:2], accessions_10k[:1]) #preview
    
    # === FILTER BY YEAR AND MAX FILINGS ===
    accessions_10q = filter_filings_by_year(accessions_10q, YEAR, N_10Q)
    accessions_10k = filter_filings_by_year(accessions_10k, YEAR, N_10K)
    
    # === If too few filings, fallback to full master index scan
    if len(accessions_10q) < N_10Q or len(accessions_10k) < N_10K:
        print(f"⚠️ Not enough filings from recent submissions — falling back to master index.")
        use_fallback = True
        log_metric("fallback_triggered", True)
    else:
        use_fallback = False
        print("✅ Using recent submissions only — fallback not needed.")
        log_metric("fallback_triggered", False)
    
    
    # In[ ]:
    
    
    # === FETCH & PARSE RECENT FILINGS ===================================
    # === Labels 10-Q accessions from EDGAR with fiscal-end dates and assign quarters ===
    
    from datetime import datetime
    
    def label_10q_accessions(accessions_10q, accessions_10k):
        
    # === Labels 10-Q accessions from EDGAR with fiscal-end dates and assign quarters ===
        """
        Assigns fiscal year-end, quarter, and label metadata to a list of 10-Q accessions 
        using the company's known 10-K fiscal year-end dates.
    
        This function matches each 10-Q's 'report_date' (document period end) to the closest 
        fiscal year-end from available 10-Ks. It then calculates the quarter number (Q1, Q2, Q3) 
        based on the number of days between the report date and fiscal year-end. Finally, it assigns:
          - 'fiscal_year_end' (YYYY-MM-DD)
          - 'quarter' (e.g., "Q2")
          - 'label' (e.g., "2Q24")
          - 'year' (based on the 10-Q's report date)
    
        This metadata is used to pre-classify filings before downloading or parsing their contents,
        enabling faster and more selective downstream processing.
    
        Args:
            accessions_10q (list of dict): List of 10-Q filings, each with a 'report_date'.
            accessions_10k (list of dict): List of 10-K filings, each with a 'report_date'.
    
        Returns:
            list of dict: Enriched list of 10-Q filings with 'fiscal_year_end', 'quarter', 'label', and 'year'.
    
        Example:
            label_10q_accessions(accessions_10q, accessions_10k)
            >> accessions_10q[0]['label'] → "3Q24"
            >> accessions_10q[0]['fiscal_year_end'] → "2024-12-31"
        """  
    
        # === Extract and sort valid fiscal year-end dates from 10-Ks ===
        fiscal_year_ends = []
        
        for entry in accessions_10k:
            fy_date = parse_date(entry["report_date"])
            if fy_date:
                fiscal_year_ends.append(fy_date)
        
        fiscal_year_ends = sorted(fiscal_year_ends, reverse=True)
        
        if not fiscal_year_ends:
            raise ValueError("No valid fiscal year-end dates found in 10-Ks.")
        
        # === Match each 10-Q to its fiscal year and assign quarter label ===
        print("\n📊 Matching 10-Qs to fiscal year-end and labeling quarters based off report date:")
        
        for q in accessions_10q:
            q_date = parse_date(q["report_date"])
            if not q_date:
                q["quarter"] = None
                q["label"] = None
                continue
        
            # Match to the fiscal year that this 10-Q falls into — first fiscal year-end after Q end
            
            # Prefer fiscal year-ends >= Q date (standard case)
            candidates = [fy for fy in fiscal_year_ends if fy >= q_date]
            
            if candidates:
                matched_fy = min(candidates)
                used_fallback = False
            else:
                # Fallback: use latest fiscal year-end before Q date
                candidates = [fy for fy in fiscal_year_ends if fy < q_date]
                matched_fy = max(candidates) if candidates else None
                used_fallback = True
            
            # 🧠 Shift forward if using fallback (e.g., using FY23 to label FY24)
            if matched_fy and used_fallback:
                matched_fy = matched_fy.replace(year=matched_fy.year + 1)
        
            if not matched_fy:
                print(f"⚠️ No matching fiscal year-end found for 10-Q ending {q['report_date']}")
                q["quarter"] = None
                q["label"] = None
                continue
        
            # Use day-based logic to assign correct quarter
            days_diff = (matched_fy - q_date).days
        
            if 70 <= days_diff <= 120:
                quarter = "Q3"
            elif 160 <= days_diff <= 200:
                quarter = "Q2"
            elif 250 <= days_diff <= 300:
                quarter = "Q1"
            else:
                print(f"⚠️ Unexpected delta ({days_diff} days) between {matched_fy.strftime('%Y-%m-%d')} and {q_date.strftime('%Y-%m-%d')} — nonstandard quarter")
                q["quarter"] = None
                q["label"] = None
                q["non_standard_period"] = True
                continue
        
            # Apply labels to the quarterly filings
            q["fiscal_year_end"] = matched_fy
            q["quarter"] = quarter
            q["calendar_year"] = q_date.year    #Note: Calendar year NOT fiscal year
            q["label"] = f"{quarter[1:]}Q{str(matched_fy.year)[-2:]}"  # e.g. "Q1" + "25" → "1Q25" #Uses fiscal year match
        
            print(f"✅ {q['report_date']} → {q['label']} (matched FY end {q['fiscal_year_end']})")
            
        return accessions_10q
    
    # === LABEL 10Q ACCESSIONS ===
    if not use_fallback:
        accessions_10q = label_10q_accessions(accessions_10q, accessions_10k)
    
    
    # In[ ]:
    
    
    # === FETCH & PARSE RECENT FILINGS ===================================
    # === Filter for required 10-Q's for workflow ===================================
    
    def filter_10q_accessions(accessions_10q, fiscal_year, quarter):
        """
        Filters 10-Q accessions based on fiscal year and quarter for quarterly workflows.
    
        This function uses the 'quarter' and 'fiscal_year_end' fields assigned during labeling,
        and extracts the fiscal year from 'fiscal_year_end' for matching.
    
        Parameters:
            accessions_10q (list of dict): List of labeled 10-Q filings
            fiscal_year (int): Target fiscal year (e.g., 2025)
            quarter (int): Target fiscal quarter number (1–4)
    
        Returns:
            list of dict: Filtered 10-Q filings needed for processing
        """
    
        # === Build list of (quarter, fiscal_year) targets ===
        targets = []
    
        if quarter == 4:
            # Q3 and Q2 of current and prior fiscal years
            for q in [3, 2]:
                targets.append((f"Q{q}", fiscal_year))
                targets.append((f"Q{q}", fiscal_year - 1))
                
        else:
            # Target quarter
            targets.append((f"Q{quarter}", fiscal_year))
        
            # Prior quarter
            if quarter > 1:
                targets.append((f"Q{quarter - 1}", fiscal_year))
            else:
                targets.append(("Q4", fiscal_year - 1))
        
            # YoY same quarter
            targets.append((f"Q{quarter}", fiscal_year - 1))
        
            # YoY prior quarter
            if quarter > 1:
                targets.append((f"Q{quarter - 1}", fiscal_year - 1))
            else:
                targets.append(("Q4", fiscal_year - 2))
    
        # === Filter using parsed fiscal year from fiscal_year_end
        filtered = [
            q for q in accessions_10q
            if (
                q.get("quarter") in ["Q1", "Q2", "Q3"] and
                q.get("fiscal_year_end") and
                (q["quarter"], q["fiscal_year_end"].year) in targets
            )
        ]
    
        print(f"✅ Selected {len(filtered)} 10-Q filings for processing.")
        return filtered
        
    # === FILTER FOR REQUIRED 10Q ACCESSIONS ===
    if not use_fallback:
        required_10q_filings = filter_10q_accessions(accessions_10q, YEAR, QUARTER)
    
    
    # In[ ]:
    
    
    # === FETCH & PARSE RECENT FILINGS ===================================
    # === Label 10K's with fiscal year data ===================================
    
    def enrich_10k_accessions_with_fiscal_year(accessions_10k):
    # === Enrich 10-K results with fiscal year metadata ===
    # Gathers fiscal year end date from recent 10-K filings
        
        """
        Enriches a list of 10-K accessions with fiscal metadata based on their report dates.
        
        For each 10-K, this function parses the 'report_date' (or 'document_period_end') to assign:
          - 'year': The fiscal year (YYYY)
          - 'fiscal_year_end': The formatted fiscal year-end date (e.g., 'December 31')
        
        This metadata enables accurate quarter matching and fiscal alignment without requiring
        full .htm downloads or XBRL parsing.
        
        Args:
            accessions_10k (list of dict): A list of 10-K filings, each containing at least 'report_date'
                                           (from SEC JSON) or 'document_period_end' (from enriched filings).
        
        Returns:
            list of dict: The enriched accessions list, with 'year' and 'fiscal_year_end' fields added.
        
        Example:
            enrich_10k_accessions_with_fiscal_year(accessions_10k)
            >> accessions_10k[0]['fiscal_year_end'] → "December 31"
            >> accessions_10k[0]['year'] → 2023
        """
    
        print("\n🛠 Enriching 10-Ks with fiscal year and fiscal year-end...")
        
        for k in accessions_10k:
            period_end = k.get("report_date")
            dt = parse_date(period_end)
        
            if dt:
                k["year"] = dt.year # note this is FISCAL year
                k["fiscal_year_end"] = dt
                print(f"✅ {period_end} → Fiscal Year {k['year']} | Fiscal Year End: {k['fiscal_year_end']}")
            else:
                k["year"] = None
                k["fiscal_year_end"] = None
                print(f"⚠️ Could not parse period end for accession {k['accession']}")
                
        return accessions_10k
    
    # === ENRICH 10K ACCESSIONS WITH FISCAL YEAR METADATA ===
    if not use_fallback:
        accessions_10k = enrich_10k_accessions_with_fiscal_year(accessions_10k)
    
    
    # In[ ]:
    
    
    # === FETCH & PARSE RECENT FILINGS ===================================
    # === Filter for required 10-K filings ===================================
    
    def filter_10k_accessions(accessions_10k, fiscal_year, quarter):
        """
        Filters 10-K accessions for workflows that require full-year and YoY start-date rollforward.
    
        For 4Q workflows, returns 10-Ks for:
          - current fiscal year
          - prior fiscal year
          - prior-prior fiscal year
    
        For Q1–Q3, returns an empty list (10-Ks not used).
    
        Parameters:
            accessions_10k (list of dict): List of 10-K filings with 'fiscal_year_end' (e.g. "2025-01-26")
            fiscal_year (int): Target fiscal year (e.g., 2025)
            quarter (int): Target fiscal quarter (1–4)
    
        Returns:
            list of dict: Filtered 10-Ks needed for processing
        """
    
        if quarter == 4:
            needed_years = {fiscal_year, fiscal_year - 1, fiscal_year - 2}
        else:
            needed_years = {fiscal_year - 1, fiscal_year - 2} #Required 10-K's for quarterly workflow
    
        filtered = [
            k for k in accessions_10k
            if (
                k["year"] in needed_years
            )
        ]
    
        print(f"✅ Selected {len(filtered)} 10-K filings for processing.")
        return filtered
    
    # === FILTER FOR REQUIRED 10Q ACCESSIONS ===
    if not use_fallback:
        required_10k_filings = filter_10k_accessions(accessions_10k, YEAR, QUARTER)
    
    
    # In[ ]:
    
    
    # === FALLBACK: FETCH & PARSE FILINGS ===================================
    # Looks into the master index to find the 10K's and 10Q's and puts them in a list
    
    import requests
    import gzip
    from io import BytesIO
    from datetime import datetime
    
    # === Calculate required year window for master index lookup
    YEARS_TO_PULL = N_10K  # Based on N_10K = 4 and N_10Q = 12
    years_to_check = list(range(YEAR - (YEARS_TO_PULL - 1), YEAR + 2))  # [2020, 2021, 2022, 2023, 2024] for 2023
    quarters_to_check = ["QTR1", "QTR2", "QTR3", "QTR4"]
    
    # === Fetch 10-Q and 10-K accessions from EDGAR master index ===
    def fetch_10q_10k_accessions_from_master (cik, headers, years=None, quarters=None):
        
        """
        Retrieves metadata for 10-Q and 10-K filings for a specified company from the SEC EDGAR master index.
    
        This function downloads and parses the quarterly master index files published by the SEC,
        filters for the given CIK and filing types ("10-Q" and "10-K"), and returns two lists:
        one for 10-Q filings and one for 10-K filings. Each entry includes the accession number,
        report date, and form type.
    
        Parameters:
            cik (str): The SEC Central Index Key (CIK) of the company, as a string. Leading zeros are optional.
            headers (dict): HTTP headers that must include a valid 'User-Agent' for SEC compliance.
            years (list of int): The list of years to check (e.g., [2022, 2023]).
            quarters (list of str): The list of quarters to check (e.g., ["QTR1", "QTR2"]).
    
        Returns:
            tuple: Two lists:
                - accessions_10q (list of dict): Each dict contains 'accession', 'report_date', and 'form' for a 10-Q.
                - accessions_10k (list of dict): Each dict contains 'accession', 'report_date', and 'form' for a 10-K.
    
        Example call:
            fetch_10q_10k_accessions_from_master("0000320193", headers, years=[2023], quarters=["QTR1", "QTR2"])
        """
        
        cik_str = str(cik).lstrip("0") # normalize
        
        accessions_10q = []
        accessions_10k = []
    
        for year in years:
            for qtr in quarters:
                url = f"https://www.sec.gov/Archives/edgar/full-index/{year}/{qtr}/master.gz"
                print(f"📦 Downloading: {url}")
                try:
                    r = requests.get(url, headers=headers)
                    time.sleep(REQUEST_DELAY)
                    r.raise_for_status()
                except Exception as e:
                    print(f"❌ Failed to fetch {year} {qtr}: {e}")
                    continue
    
                # Decompress and decode
                with gzip.open(BytesIO(r.content), 'rt', encoding='latin-1') as f:
                    started = False
                    for line in f:
                        if not started:
                            if line.strip().startswith("CIK|"):
                                started = True
                            continue
    
                        parts = line.strip().split("|")
                        if len(parts) != 5:
                            continue
    
                        cik_field, company, form, date_filed, filename = parts
    
                        if cik_field != cik_str:
                            continue  # skip other companies
    
                        if form not in ("10-Q", "10-K"):
                            continue  # skip other forms
    
                        accession = filename.split("/")[-1].replace(".txt", "")
                        entry = {
                            "accession": accession,
                            "report_date": date_filed, #this is the filing date - but using report_date preserve logic downstream
                            "form": form
                        }
    
                        if form == "10-Q":
                            accessions_10q.append(entry)
                        elif form == "10-K":
                            accessions_10k.append(entry)
    
        print(f"✅ Found {len(accessions_10q)} 10-Q accessions (from master index)")
        print(f"✅ Found {len(accessions_10k)} 10-K accessions (from master index)")
        return accessions_10q, accessions_10k
    
    # === FETCH 10Q/10K ACCESSIONS FROM MASTER INDEX ===
    
    if len(accessions_10q) < N_10Q or len(accessions_10k) < N_10K:
        accessions_10q, accessions_10k = fetch_10q_10k_accessions_from_master(CIK, HEADERS, years_to_check, quarters_to_check)
        print(accessions_10q[:2], accessions_10k[:1])
    
    
    # In[ ]:
    
    
    # === FETCH & PARSE FILINGS ===================================
    # Extracts data from filings starting from the target filing from CONFIG to use
    # === Extract facts from 10-Q and 10-K accessions from EDGAR ===
    
    import time
    start_total = time.time()
    
    from bs4 import BeautifulSoup
    from bs4 import XMLParsedAsHTMLWarning
    import warnings
    warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
    
    # === CONFIG ===
    STOP_AFTER_FIRST_VALID_PERIOD = True
    
    # === Extract facts and DocumentPeriodEndDate from a single .htm ===
    def extract_facts_with_document_period(ixbrl_url, headers):
    
        """
        Parses a single EDGAR inline XBRL (.htm) filing and extracts all tagged financial facts,
        associated XBRL context blocks, and the filing's document period end date.
    
        Args:
            ixbrl_url (str): Full URL to the .htm iXBRL filing on EDGAR.
            headers (dict): HTTP headers to use in the request (User-Agent required by SEC).
    
        Returns:
            dict: A dictionary with the following keys:
                - 'facts': List of extracted facts (each as dict with tag, contextref, value, text)
                - 'context_blocks': Dict mapping contextRef IDs to raw XBRL context HTML
                - 'document_period_end': Extracted DEI tag (str) for DocumentPeriodEndDate
                - 'document_period_label': Raw label text of the DocumentPeriodEndDate (if available)
    
        Example:
            extract_facts_with_document_period("https://www.sec.gov/Archives/...", headers)
        """
        print(f"\n🌐 Fetching iXBRL: {ixbrl_url}")
        t0 = time.time()
    
        r = requests.get(ixbrl_url, headers=headers)
        fetch_time = time.time() - t0
        print(f"⏳ Fetch time: {fetch_time:.2f} seconds")
    
        time.sleep(REQUEST_DELAY)
        r.raise_for_status()
    
        t1 = time.time()
        soup = BeautifulSoup(r.content, "lxml")
        parse_time = time.time() - t1
        print(f"🧠 Parse time: {parse_time:.2f} seconds")
    
        # === Dynamic slowdown warning ===
        content_mb = len(r.content) / 1_000_000
        if content_mb > 3:
            print(f"⚠️ Large filing ({content_mb:.1f} MB) — this may take a minute...")
    
        ix_tags = soup.find_all(["ix:nonfraction", "ix:nonnumeric"])
        print(f"📦 Found {len(ix_tags)} ix: tags")
        
        if len(ix_tags) > 800:
            print(f"⚠️ Detected {len(ix_tags)} facts — parsing may take a minute...")
        
        facts = []
        context_blocks = {}  # 🆕 New dictionary to store contexts
        doc_period_end = None
        doc_period_label = None
    
        # --- First: Extract all context blocks ---
        for ctx_tag in soup.find_all("xbrli:context"):
            ctx_id = ctx_tag.get("id")
            if ctx_id:
                context_blocks[ctx_id] = str(ctx_tag)  # Save the raw HTML block
    
        # --- Then: Extract all facts ---
        for tag in soup.find_all(["ix:nonfraction", "ix:nonnumeric"]):
            name = tag.get("name")
            ctx = tag.get("contextref")
            sign = tag.get("sign")
            val = tag.text or tag.get("value") or "".join(tag.stripped_strings)
    
            if not (name and ctx and val):
                continue
    
            if name == "dei:DocumentPeriodEndDate":
                doc_period_end = val.strip()
                doc_period_label = tag.text.strip()
    
            try:
                value = float(val.replace(",", "").replace("−", "-"))
                if sign == "-":  # ✅ New: apply sign flip
                    value = -abs(value)
            except ValueError:
                continue
    
            facts.append({
                "tag": name,
                "contextref": ctx,
                "value": value,
                "text": tag.text.strip()
            })
    
        return {
            "facts": facts,
            "context_blocks": context_blocks,  # 🆕 Include context_blocks in the return!
            "document_period_end": doc_period_end,
            "document_period_label": doc_period_label
        }
    
    # === Try all .htm files inside an accession (starts with largest file, then stop after first valid) ===
    def try_all_htm_files(cik, accession_number, headers):
    
        """
        Attempts to extract financial data from .htm files in a given EDGAR accession by scanning for 
        the first valid inline XBRL (iXBRL) filing. Starts with the largest .htm file by size, then 
        falls back to other .htm files if needed.
    
        Args:
            cik (str or int): Central Index Key (CIK) of the company.
            accession_number (str): SEC accession number (e.g., "0000320193-23-000055").
            headers (dict): HTTP headers for SEC requests (must include User-Agent).
    
        Returns:
            list of dict: A list containing one or more parsed iXBRL results, each with:
                - 'file': Filename of the .htm file parsed
                - 'url': Full URL to the .htm file
                - 'document_period_end': DEI DocumentPeriodEndDate (str)
                - 'document_period_label': Human-readable label of the period
                - 'facts': List of extracted financial fact dictionaries
                - 'context_blocks': Dict of raw XBRL context XML blocks
                - 'concept_roles': Mapping of tags to presentation roles from .pre.xml
    
        Behavior:
            - Prioritizes the largest .htm file by size, assuming it is the main filing.
            - Requires at least 50 extracted XBRL facts to consider a file valid.
            - Stops after the first valid file unless STOP_AFTER_FIRST_VALID_PERIOD is False.
            - Falls back to other .htm files in the accession directory if the largest file is invalid.
    
        Example:
            results = try_all_htm_files("320193", "0000320193-23-000055", headers)
        """
    
        acc_nodash = accession_number.replace("-", "")
        index_url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc_nodash}/index.json"
        base_url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc_nodash}/"
        
        try:
            r = requests.get(index_url, headers=headers)
            time.sleep(REQUEST_DELAY)
            r.raise_for_status()
            index = r.json()
        except Exception as e:
            print(f"❌ Failed to fetch index.json for {accession_number}: {e}")
            return []
    
        items = index.get("directory", {}).get("item", [])
        results = []
        
        # === Try largest .htm file by size first ===    
        htm_items = [
            item for item in items
            if item["name"].lower().endswith(".htm") and item.get("size", "").isdigit()
        ]
        
        if htm_items:
            # Sort descending by file size
            htm_items.sort(key=lambda x: int(x["size"]), reverse=True)
            largest_htm = htm_items[0]["name"]
            full_url = base_url + largest_htm
            print(f"📏 Trying largest .htm file first: {largest_htm} ({htm_items[0]['size']} bytes)")
    
            try:
                data = extract_facts_with_document_period(full_url, headers)
                
                if data["document_period_end"] and len(data["facts"]) >= 50:
                    # Fetch presentation roles from .pre.xml                    
                    concept_roles = get_concept_roles_from_presentation(cik, accession_number, headers)
                    
                    print(f"✅ {full_url} → Period End: {data['document_period_end']}")
                    print(f"🔎 Extracted {len(data['facts'])} facts")
                
                    results.append({
                        "file": largest_htm,
                        "url": full_url,
                        "document_period_end": data["document_period_end"],
                        "document_period_label": data["document_period_label"],
                        "facts": data["facts"],
                        "context_blocks": data["context_blocks"],
                        "concept_roles": concept_roles
                    })
                    return results  # ✅ Success: stop here
                                        
            except Exception as e:
                print(f"⚠️ Error checking largest .htm file: {e}")
    
        # === Fallback: Try all other .htm files ===
        print("🔁 Fallback: checking all .htm files...")
        for item in items:
            name = item["name"].lower()
            if not name.endswith(".htm"):
                continue
    
            full_url = base_url + item["name"]
            try:
                data = extract_facts_with_document_period(full_url, headers)
    
                if data["document_period_end"]:
                    if len(data["facts"]) < 50:
                        print(f"⚠️ Warning: only {len(data['facts'])} facts extracted from {full_url} — possible exhibit or junk file.")
                        continue  # Skip this .htm and keep looking
    
                    # Fetch presentation roles from .pre.xml
                    concept_roles = get_concept_roles_from_presentation(cik, accession_number, headers)
                    
                    print(f"✅ {full_url} → Period End: {data['document_period_end']}")
                    print(f"🔎 Extracted {len(data['facts'])} facts")
                          
                    results.append({
                        "file": item["name"],
                        "url": full_url,
                        "document_period_end": data["document_period_end"],
                        "document_period_label": data["document_period_label"],
                        "facts": data["facts"],
                        "context_blocks": data["context_blocks"],  # 🆕 Capture context blocks too
                        "concept_roles": concept_roles
                    })
                    if STOP_AFTER_FIRST_VALID_PERIOD:
                        break  # 🔥 Stop scanning more .htms once one good file is found
    
            except Exception as e:
                print(f"⚠️ Error checking {item['name']}: {e}")
                continue
    
        return results
        
    # === Extract information from list of 10K and 10Q filings ===
    def extract_filing_batch(accessions, cik, headers, form_type):
    
        """
        Processes a batch of 10-Q or 10-K filings and extracts financial data from each valid .htm file.
    
        Args:
            accessions (list of dict): List of accessions with keys 'accession' and 'report_date'.
            cik (str or int): Central Index Key for the company.
            headers (dict): HTTP headers to use for SEC requests.
            form_type (str): Filing type to label results (e.g., "10-Q" or "10-K").
    
        Returns:
            list of dict: One result per valid filing, with keys:
                - 'accession': Accession number
                - 'report_date': Filing or report date
                - 'file': .htm filename
                - 'url': Full .htm URL
                - 'document_period_end': DEI period end date
                - 'document_period_label': Raw label for the period
                - 'facts': Extracted financial facts
                - 'context_blocks': XBRL contexts
                - 'concept_roles': Extracted presentation role labels
                - 'form': Filing form type
    
        Example:
            extract_filing_batch(accessions_10q, "320193", headers, "10-Q")
        """
        
        results = []
        for i, entry in enumerate(accessions):
            acc = entry["accession"]
            report_date = entry["report_date"]
    
            # 🚫 Skip filings before 2019 (no inline XBRL guaranteed)
            if int(report_date[:4]) < 2019:
                print(f"⏩ Skipping {acc} — pre-2019 filing")
                continue
                
            print(f"\n🔍 {form_type} Accession {i+1}: {acc} | Report or Filing Date: {report_date}")
            extracted = try_all_htm_files(cik, acc, headers)
            
            if not extracted:
                continue
            for result in extracted:
                results.append({
                    "accession": acc,
                    "report_date": report_date,
                    "file": result["file"],
                    "url": result["url"],
                    "document_period_end": result["document_period_end"],
                    "document_period_label": result["document_period_label"],
                    "facts": result["facts"],
                    "context_blocks": result["context_blocks"],
                    "concept_roles": result["concept_roles"],
                    "form": form_type
                })
        return results
    
    # === EXTRACT INFORMATION FROM 10-Qs and 10-K's ===
    
    if not use_fallback:
        print("\n📘 Processing 10-Qs...")
        results_10q = extract_filing_batch(required_10q_filings, CIK, HEADERS, "10-Q")
    
        print("\n📕 Processing 10-Ks...")
        results_10k = extract_filing_batch(required_10k_filings, CIK, HEADERS, "10-K")
        
    else:    
        print("\n⚠️ Skipping filtered extraction — fallback mode will parse full lists.")
        
        print("\n📘 Processing 10-Qs...")
        results_10q = extract_filing_batch(accessions_10q, CIK, HEADERS, "10-Q")
        
        print("\n📕 Processing 10-Ks...")
        results_10k = extract_filing_batch(accessions_10k, CIK, HEADERS, form_type="10-K")
    
    # === CALCULATING PROCESSING TIME ===
    
    print("🎯 All filings processed. Script complete!")
    end_total = time.time()
    print(f"\n⏱️ Total extraction processing time: {end_total - start_total:.2f} seconds")
    log_metric("extraction_processing_seconds", round(end_total - start_total, 2))
    
    
    # In[ ]:
    
    
    # === FETCH & PARSE FILINGS ===================================
    # === Labels 10-Q accessions from EDGAR with fiscal-end dates and assign quarters ===
    
    from datetime import datetime
    
    # === Extract and sort valid fiscal year-end dates from 10-Ks ===
    fiscal_year_ends = []
    
    for entry in results_10k:
        fy_date = parse_date(entry["document_period_end"])
        if fy_date:
            fiscal_year_ends.append(fy_date)
    
    fiscal_year_ends = sorted(fiscal_year_ends, reverse=True)
    
    if not fiscal_year_ends:
        raise ValueError("No valid fiscal year-end dates found in 10-Ks.")
    
    # === Match each 10-Q to its fiscal year and assign quarter label ===
    print("\n📊 Matching 10-Qs to fiscal year-end and labeling quarters based off report date:")
    
    for q in results_10q:
        q_date = parse_date(q["document_period_end"])
        if not q_date:
            q["quarter"] = None
            q["label"] = None
            continue
    
        # Match to the fiscal year that this 10-Q falls into — first fiscal year-end after Q end
        
        # Prefer fiscal year-ends >= Q date (standard case)
        candidates = [fy for fy in fiscal_year_ends if fy >= q_date]
        
        if candidates:
            matched_fy = min(candidates)
            used_fallback = False
        else:
            # Fallback: use latest fiscal year-end before Q date
            candidates = [fy for fy in fiscal_year_ends if fy < q_date]
            matched_fy = max(candidates) if candidates else None
            used_fallback = True
        
        # 🧠 Shift forward if using fallback (e.g., using FY23 to label FY24)
        if matched_fy and used_fallback:
            matched_fy = matched_fy.replace(year=matched_fy.year + 1)
    
        if not matched_fy:
            print(f"⚠️ No matching fiscal year-end found for 10-Q ending {q['document_period_end']}")
            q["quarter"] = None
            q["label"] = None
            continue
    
        # Use day-based logic to assign correct quarter
        days_diff = (matched_fy - q_date).days
    
        if 70 <= days_diff <= 120:
            quarter = "Q3"
        elif 160 <= days_diff <= 200:
            quarter = "Q2"
        elif 250 <= days_diff <= 300:
            quarter = "Q1"
        else:
            print(f"⚠️ Unexpected delta ({days_diff} days) between {matched_fy.strftime('%Y-%m-%d')} and {q_date.strftime('%Y-%m-%d')} — nonstandard quarter")
            q["quarter"] = None
            q["label"] = None
            q["non_standard_period"] = True
            continue
    
        # Apply labels to the quarterly filings
        q["fiscal_year_end"] = matched_fy
        q["quarter"] = quarter
        q["year"] = q_date.year    #Note: Fiscal year
        q["label"] = f"{quarter[1:]}Q{str(matched_fy.year)[-2:]}"  # e.g. "Q1" + "25" → "1Q25" #Uses fiscal year match
    
        print(f"✅ {q['document_period_end']} → {q['label']} (matched FY end {q['fiscal_year_end']})")
    
    
    # In[ ]:
    
    
    # === FETCH & PARSE FILINGS ===================================
    # Gathers fiscal year end date from recent 10-K filings
    # === Enrich 10-K results with fiscal year metadata ===
    
    print("\n🛠 Enriching 10-Ks with fiscal year and fiscal year-end...")
    
    for k in results_10k:
        period_end = k.get("document_period_end")
        dt = parse_date(period_end)
    
        if dt:
            k["year"] = dt.year
            k["fiscal_year_end"] = dt
            print(f"✅ {period_end} → Fiscal Year {k['year']} | Fiscal Year End: {k['fiscal_year_end']}")
        else:
            k["year"] = None
            k["fiscal_year_end"] = None
            print(f"⚠️ Could not parse period end for accession {k['accession']}")
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    #Identifies the target 10-Q filing to use
    # === Filter 10-Qs Based on YEAR and QUARTER from CONFIG for Target Filings ===
    
    # Build target label from your existing config
    target_label = f"{QUARTER}Q{str(YEAR)[-2:]}"  # e.g., 2Q24
    print(f"\n🎯 Target Label: {target_label}")
    
    if FOUR_Q_MODE:
        print("📄 4Q mode detected: will select 10-K filing and prior 10-Q's instead of specific 10-Q.")
        target_10q = None
      
    else:
        # Normal flow: Pick 10-Q
        # === Filter results_10q based on target label
        filtered_10qs = [q for q in results_10q if q.get("label") == target_label]
        
        # === Output matching 10-Qs
        print(f"\n📄 Matching 10-Qs for {target_label}:")
        
        if not filtered_10qs:
            target_10q = None
            print(f"⚠️ No 10-Q filings found for {target_label}.")
        else:
            target_10q = filtered_10qs[0]
            for q in filtered_10qs:
                print(f"✅ {q['label']} | Period End: {q['document_period_end']} | URL: {q['url']}")
    
    # === Log Target 10-Q ===
    if not FOUR_Q_MODE and target_10q:
        log_metric("target_filing", {
            "type": "10-Q",
            "label": target_10q["label"],
            "period_end": target_10q["document_period_end"],
            "url": target_10q["url"]
        })
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # Identifies the target 10-K and prior 10-K, and current year Q3 10-Q
    
    # Build annual label from your existing config
    annual_label = f"FY{str(YEAR)[-2:]}"  # Example: "FY24"
    print(f"\n🎯 Annual Label: {annual_label}")
    
    if FOUR_Q_MODE:
    # Select current year 10-K
        filtered_10ks = [k for k in results_10k if k.get("year") == YEAR]
        if not filtered_10ks:
            raise ValueError(f"❌ No matching 10-K found for {YEAR}.")
        target_10k = filtered_10ks[0]
        print(f"Selected 10-K for full year: Period-End: {target_10k['document_period_end']}")
        print(f"URL: {target_10k['url']}")
    
        # Use fiscal year end from selected 10-K. 
        # Note: fiscal_year_end is in YYYY-MM-DD string format (assigned during quarter labeling step)
        
        fye_target = target_10k["fiscal_year_end"]
        
        # Select current year Q1–Q3 10-Qs by fiscal year end
        q1_entry = next((q for q in results_10q if q.get("quarter") == "Q1" and q.get("fiscal_year_end") == fye_target), None)
        q2_entry = next((q for q in results_10q if q.get("quarter") == "Q2" and q.get("fiscal_year_end") == fye_target), None)
        q3_entry = next((q for q in results_10q if q.get("quarter") == "Q3" and q.get("fiscal_year_end") == fye_target), None)
    
        # === Store Quarter Entries in a Dict ===
        quarter_entries = {
            "Q1": q1_entry,
            "Q2": q2_entry,
            "Q3": q3_entry
        }
    
        # Check for missing
        missing_qs = []
        if not q1_entry: missing_qs.append("Q1")
        if not q2_entry: missing_qs.append("Q2")
        if not q3_entry: missing_qs.append("Q3")
    
        if missing_qs:
            print(f"\n⚠️ Missing current year 10-Qs for: {', '.join(missing_qs)}")
            if "Q1" in missing_qs: q1_entry = None
            if "Q2" in missing_qs: q2_entry = None
            if "Q3" in missing_qs: q3_entry = None
    
        print(f"\n✅ Found Q1-Q3 10-Qs for fiscal year {YEAR}:")
    
        for q, entry in quarter_entries.items():
            if entry and "document_period_end" in entry and "url" in entry:
                print(f"   -{q}: Period End: {entry['document_period_end']} | {entry['url']}")
    
        if not q3_entry:
            raise ValueError("❌ Missing current year Q3 10-Q — required for 4Q processing.")
    
        # Select prior year 10-K - this may be redundant - used in next step
        prior_10k = next((k for k in results_10k if k.get("year") == YEAR - 1), None)
        
        if not prior_10k:
            raise ValueError(f"❌ Missing prior year 10-K for {YEAR - 1} — required for prior 4Q calculation.")
    
        print(f"\n✅ Selected prior 10-K: Period-End: {prior_10k['document_period_end']}")
        print(f"URL: {prior_10k['url']}")
            
        # === Log Target 10-K ===
        log_metric("target_filing", {
            "type": "10-K",
            "label": annual_label,
            "period_end": target_10k["document_period_end"],
            "url": target_10k["url"]
        })   
    
    else:
        target_10k = None
        q1_entry = None
        q2_entry = None
        q3_entry = None
        prior_10k = None
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW and 4Q WORKFLOW =============================================
    # Identifies the prior quarterly and annual filings to use 
    
    # === Normal mode (e.g., 1Q, 2Q, or 3Q)
    # === STEP: Find prior y/y Q filings from results_10q
    
    if not FOUR_Q_MODE:
    
        quarter = target_10q.get("quarter")
        fye_str = target_10q.get("fiscal_year_end")  # already in 'YYYY-MM-DD' format
    
        # Identify previous fiscal year-end from known values (sorted descending)
        fiscal_ends = sorted({q.get("fiscal_year_end") for q in results_10q if q.get("fiscal_year_end")}, reverse=True)
        try:
            idx = fiscal_ends.index(fye_str)
            prior_fye_str = fiscal_ends[idx + 1]  # next fiscal year end in time
        except (ValueError, IndexError):
            prior_fye_str = None
    
        prior_10q = None
        if prior_fye_str:
            prior_10q = next(
                (q for q in results_10q if q.get("quarter") == quarter and q.get("fiscal_year_end") == prior_fye_str),
                None
            )
        
        if prior_10q:
            print(f"\n✅ Found prior 10-Q: {prior_10q['label']}")
            print(f"Period End: {prior_10q['document_period_end']}")
            print(f"URL: {prior_10q['url']}")
        else:  
            print(f"\n⚠️ Could not find prior 10-Q.")
    
    else:
        
    # === 4Q mode: find prior 10-K and prior Q3 10-Qs
        prior_10k = next((k for k in results_10k if k.get("year") == (YEAR - 1)), None)
        
        if prior_10k:
            print(f"\n✅ Found prior 10-K for {YEAR-1}:")
            print(f"Period End: {prior_10k['document_period_end']}")
            print(f"URL: {prior_10k['url']}")
        else:
            print(f"\n⚠️ Could not find prior 10-K.")
    
        # Use the prior 10-K document period end as fiscal year end anchor
        fye_prior = prior_10k["fiscal_year_end"] if prior_10k else None
    
        # Match prior 10-Qs with the same fiscal year end and correct quarter
        q1_prior_entry = next((q for q in results_10q if q.get("quarter") == "Q1" and q.get("fiscal_year_end") == fye_prior), None)
        q2_prior_entry = next((q for q in results_10q if q.get("quarter") == "Q2" and q.get("fiscal_year_end") == fye_prior), None)
        q3_prior_entry = next((q for q in results_10q if q.get("quarter") == "Q3" and q.get("fiscal_year_end") == fye_prior), None)
    
        # === Store Prior Quarter Entries in a Dict ===
        prior_quarter_entries = {
            "Q1": q1_prior_entry,
            "Q2": q2_prior_entry,
            "Q3": q3_prior_entry
        }
    
        if not FULL_YEAR_MODE:
            print(f"\n✅ Prior 10-Qs found for fiscal year {YEAR - 1}:")
    
        for q_label, q_entry in prior_quarter_entries.items():
            if q_entry:
                print(f"  - {q_label}: Period End: {q_entry['document_period_end']} | URL: {q_entry['url']}")
    
        if not q3_prior_entry:
            raise ValueError("❌ Missing prior year Q3 10-Q — required for 4Q processing.")
    
        # Check for missing entries
        missing_prior_qs = []
        if not q1_prior_entry: missing_prior_qs.append("Q1")
        if not q2_prior_entry: missing_prior_qs.append("Q2")
        if not q3_prior_entry: missing_prior_qs.append("Q3")
    
        if missing_prior_qs:
            print(f"\n⚠️ Missing current year 10-Qs for: {', '.join(missing_prior_qs)}")
            if "Q1" in missing_prior_qs: q1_prior_entry = None
            if "Q2" in missing_prior_qs: q2_prior_entry = None
            if "Q3" in missing_prior_qs: q3_prior_entry = None
    
        if FULL_YEAR_MODE:
            print("⚠️ Skipping prior Q1–Q3 10-Q check — not needed in full-year mode.")
    
    
    # In[ ]:
    
    
    # === SHARED LOGIC (e.g. negated labels, exports) =============
    # === Get Negated Labels (for Visual Presentation) ===
    
    if FOUR_Q_MODE:
        if target_10k is None:
            raise ValueError("❌ target_10k is None — check 10-K selection.")
        negated_tags = get_negated_label_concepts(CIK, target_10k["accession"], HEADERS)
    
    else:
        if target_10q is None:
            raise ValueError("❌ target_10q is None — check 10-Q selection.")
        negated_tags = get_negated_label_concepts(CIK, target_10q["accession"], HEADERS)
    
    
    # In[ ]:
    
    
    # === SHARED LOGIC (e.g. negated labels, exports) =============
    # === Extract Concept Roles from .pre.xml for mapping ===
    
    if FOUR_Q_MODE:
        if target_10k is None:
            raise ValueError("❌ target_10k is None — check 10-K selection.")
        concept_roles = get_concept_roles_from_presentation(CIK, target_10k["accession"], HEADERS)
        filename_roles_export = f"{CIK}_{annual_label}_presentation_roles.csv"
    
    else:
        if target_10q is None:
            raise ValueError("❌ target_10q is None — check 10-Q selection.")
        concept_roles = get_concept_roles_from_presentation(CIK, target_10q["accession"], HEADERS)
        filename_roles_export = f"{CIK}_{target_label}_presentation_roles.csv"
    
    # Convert to DataFrame
    rows = []
    for tag, roles in concept_roles.items():
        for role in roles:
            rows.append({"tag": tag, "presentation_role": role})
    
    df_concept_roles = pd.DataFrame(rows)
    
    # Preview
    print(f"✅ Extracted {len(df_concept_roles)} concept→role entries from .pre.xml")
    log_metric("concept_roles_extracted", len(df_concept_roles))
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # Enrich target 10-Q and prior 10-Q data with labels
     
    if not FOUR_Q_MODE:
        print("\n📥 Enriching normal 10-Q mode...")
    
        # Enrich current 10-Q
    
        df_current = enrich_filing(target_10q)
        print(df_current["matched_category"].value_counts(dropna=False))
        categorized_Q_fact_counts = df_current["matched_category"].value_counts(dropna=False).to_dict()
        
        # Enrich prior 10-Q
    
        if prior_10q is None:
            raise ValueError(f"❌ No prior 10-Q found for {TICKER} {QUARTER}Q{YEAR} — cannot proceed without prior comparison.")
        
        else:
            df_prior = enrich_filing(prior_10q)
    
        print(df_prior["matched_category"].value_counts(dropna=False))
        log_metric("fact_category_counts", categorized_Q_fact_counts)
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # Enrich target 10-K's and prior 10-Q's data with labels (for 4Q calculations)
    
    if FOUR_Q_MODE:
        print("\n📥 Enriching 4Q mode (10-K + Prior 10-Q's...)")
    
        # Current year
        df_current_10k = enrich_filing(target_10k)
        print("🔹 Current Year 10-K facts enriched:")
        print(df_current_10k["matched_category"].value_counts(dropna=False))
        categorized_K_fact_counts = df_current_10k["matched_category"].value_counts(dropna=False).to_dict()
    
        df_q1 = df_q2 = df_q3 = None
        if q1_entry:
            df_q1 = enrich_filing(q1_entry)
            print("🔹 Current Year Q1 facts enriched:")
            print(df_q1["matched_category"].value_counts(dropna=False))
    
        if q2_entry:
            df_q2 = enrich_filing(q2_entry)
            print("🔹 Current Year Q2 facts enriched:")
            print(df_q2["matched_category"].value_counts(dropna=False))
    
        if q3_entry:
            df_q3 = enrich_filing(q3_entry)
            print("🔹 Current Year Q3 facts enriched:")
            print(df_q3["matched_category"].value_counts(dropna=False))
    
        # Prior year
        df_prior_10k = enrich_filing(prior_10k)
        print("🔹 Prior Year 10-K facts enriched:")
        print(df_prior_10k["matched_category"].value_counts(dropna=False))
    
        df_q1_prior = df_q2_prior = df_q3_prior = None
        if q1_prior_entry:
            df_q1_prior = enrich_filing(q1_prior_entry)
            print("🔹 Prior Year Q1 facts enriched:")
            print(df_q1_prior["matched_category"].value_counts(dropna=False))
    
        if q2_prior_entry:
            df_q2_prior = enrich_filing(q2_prior_entry)
            print("🔹 Prior Year Q2 facts enriched:")
            print(df_q2_prior["matched_category"].value_counts(dropna=False))
    
        if q3_prior_entry:
            df_q3_prior = enrich_filing(q3_prior_entry)
            print("🔹 Prior Year Q3 facts enriched:")
            print(df_q3_prior["matched_category"].value_counts(dropna=False))
    
        log_metric("fact_category_counts", categorized_K_fact_counts)
    
    else:
        pass
    
    
    # In[ ]:
    
    
    # === SHARED LOGIC (e.g. negated labels, exports) =============
    # === Check Negated Labels ===
    
    import pandas as pd
    
    # Ensure negated_tags is a set
    negated_list = sorted(list(negated_tags))
    
    # Create a DataFrame
    df_negated_labels = pd.DataFrame(negated_list, columns=["tag_with_negated_label"])
    
    # Preview in notebook
    print(f"✅ Found {len(df_negated_labels)} tags with negated labels in .pre.xml")
    log_metric("negated_labels_extracted", len(df_negated_labels))
    
    
    # In[ ]:
    
    
    # === SHARED LOGIC (Enrichment summary) =============
    #Preview of the enrichment results for target current and prior year filing with export to review
    
    if FOUR_Q_MODE:
        print(f"📄 Extracted facts from target: {target_10k.get('form', 'Unknown')} ending {target_10k.get('document_period_end', 'Unknown')}")
        print(f"✅ Fact categorization summary (logged): {categorized_K_fact_counts}")
    
    else:
        print(f"📄 Extracted facts from target: {target_10q.get('form', 'Unknown')} ending {target_10q.get('document_period_end', 'Unknown')}")
        print(f"✅ Fact categorization summary (logged): {categorized_Q_fact_counts}")
        
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === CLEAN AXIS VALUES FIRST ===
    if FOUR_Q_MODE:
        
        for col in AXIS_COLS:
            df_current_10k[col] = df_current_10k[col].fillna("__NONE__")
            df_prior_10k[col] = df_prior_10k[col].fillna("__NONE__")
            df_q3[col] = df_q3[col].fillna("__NONE__")
            df_q3_prior[col] = df_q3_prior[col].fillna("__NONE__")
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === MATCH FY: Current FY vs Prior FY using tag + date type from current 10-K ===
    if FOUR_Q_MODE:
        
        MATCH_COLS_FY = ["tag", "date_type"] + AXIS_COLS
        MIN_MATCH_COLS_FY = ["tag", "date_type"]
        
        # Pull both from current 10-K
        df_fy_curr = df_current_10k[df_current_10k["matched_category"] == "current_full_year"].copy()
        df_fy_prior = df_current_10k[df_current_10k["matched_category"] == "prior_full_year"].copy()
    
        for col in AXIS_COLS:
            df_fy_curr[col] = df_fy_curr[col].fillna("__NONE__")
            df_fy_prior[col] = df_fy_prior[col].fillna("__NONE__")
    
        # Adaptive match
        match_keys_fy = run_adaptive_match_keys(df_fy_curr, df_fy_prior, MATCH_COLS_FY, MIN_MATCH_COLS_FY)
        df_fy_matched = zip_match_in_order(df_fy_curr, df_fy_prior, match_keys_fy)
        df_fy_matched = standardize_zip_output(df_fy_matched)
        
        # Preview results
        total_fy_curr = len(df_fy_curr)
        matched_fy_rows = len(df_fy_matched)
        match_rate_fy = matched_fy_rows / max(total_fy_curr, 1)
        
        print(f"✅ FY match rate: {match_rate_fy * 100:.1f}%")
        log_metric("match_rate", {"fy": match_rate_fy})
    
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === MATCH YTD (from Q3 10-Q in the fiscal year): current_ytd vs prior_ytd ===
    if FOUR_Q_MODE:
    
        MATCH_COLS_YTD = ["tag", "date_type"] + AXIS_COLS
        MIN_MATCH_COLS_YTD = ["tag", "date_type"]
    
        # Pull both from the same Q3 filing
        df_ytd_curr = df_q3[df_q3["matched_category"] == "current_ytd"].copy()
        df_ytd_prior = df_q3[df_q3["matched_category"] == "prior_ytd"].copy()
    
        # Ensure axis consistency
        for col in AXIS_COLS:
            df_ytd_curr[col] = df_ytd_curr[col].fillna("__NONE__")
            df_ytd_prior[col] = df_ytd_prior[col].fillna("__NONE__")
    
        # No need to shift 'end' — both from same Q3 filing
        match_keys_ytd = run_adaptive_match_keys(df_ytd_curr, df_ytd_prior, MATCH_COLS_YTD, MIN_MATCH_COLS_YTD)
        df_ytd_matched = zip_match_in_order(df_ytd_curr, df_ytd_prior, match_keys_ytd)
        df_ytd_matched = standardize_zip_output(df_ytd_matched)
    
        # Preview results
        total_ytd_curr = len(df_ytd_curr)
        matched_ytd_rows = len(df_ytd_matched)
        match_rate_ytd = matched_ytd_rows / max(total_ytd_curr, 1)
    
        print(f"✅ FY match rate: {match_rate_ytd * 100:.1f}%")
        log_metric("match_rate", {"ytd": match_rate_ytd})
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === CHECK FOR DUPLICATE MERGE KEYS AND FULL DUPLICATE ROWS ===
    
    if FOUR_Q_MODE:
    
        merge_key = ["tag", "date_type"] + AXIS_COLS
    
        # Recalculate _key
        df_fy_matched["_key"] = df_fy_matched[merge_key].apply(tuple, axis=1)
        df_ytd_matched["_key"] = df_ytd_matched[merge_key].apply(tuple, axis=1)
    
        # Count duplicate keys (rows with same _key)
        dupes_fy_keys = df_fy_matched["_key"].duplicated(keep=False).sum()
        dupes_ytd_keys = df_ytd_matched["_key"].duplicated(keep=False).sum()
    
        # Count unique keys and shared
        keys_fy = set(df_fy_matched["_key"])
        keys_ytd = set(df_ytd_matched["_key"])
        shared_keys = keys_fy & keys_ytd
    
        # Check full row duplicates
        full_dupes_fy = df_fy_matched.duplicated(keep=False).sum()
        full_dupes_ytd = df_ytd_matched.duplicated(keep=False).sum()
    
        print("🔍 Duplicate Key & Row Check:")
        print(f"  • FY matched: {dupes_fy_keys} rows with duplicate _key") 
        print(f"  • YTD matched: {dupes_ytd_keys} rows with duplicate _key")
        print(f"  • Unique keys in FY matched: {len(keys_fy)}")
        print(f"  • Unique keys in YTD matched: {len(keys_ytd)}")
        print(f"  • Shared keys between FY and YTD: {len(shared_keys)}")
        print("—")
        print(f"  • FY matched: {full_dupes_fy} fully duplicated rows")
        print(f"  • YTD matched: {full_dupes_ytd} fully duplicated rows")
    
        if FOUR_Q_MODE:
            
            # Identify rows with duplicated _key
            is_key_duplicated_fy = df_fy_matched["_key"].duplicated(keep=False)
            is_key_duplicated_ytd = df_ytd_matched["_key"].duplicated(keep=False)
        
            # Identify full-row duplicates
            is_full_dup_fy = df_fy_matched.duplicated(keep=False)
            is_full_dup_ytd = df_ytd_matched.duplicated(keep=False)
        
            # Check overlap
            key_and_full_fy = (is_key_duplicated_fy & is_full_dup_fy).sum()
            key_and_full_ytd = (is_key_duplicated_ytd & is_full_dup_ytd).sum()
        
            print("\n🔎 Duplicate Overlap Check:")
            print(f"  • FY matched: {key_and_full_fy} rows are both _key duplicates AND full duplicates")
            print(f"  • YTD matched: {key_and_full_ytd} rows are both _key duplicates AND full duplicates")
            print("\n✅ If these numbers equal the key duplicate count, it's safe to drop.")
    
            # Show rows with same _key but not fully duplicated
            non_exact_dupes_ytd = df_ytd_matched[
                df_ytd_matched["_key"].duplicated(keep=False) & ~df_ytd_matched.duplicated(keep=False)
            ]
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === CALCULATE 4Q: From Matched FY and YTD using aligned suffixes ===
    if FOUR_Q_MODE:
    
        merge_key = ["tag", "start_current"] + AXIS_COLS
    
        # Create merged key column to use for matching
        df_fy_matched["_key"] = df_fy_matched[merge_key].apply(tuple, axis=1)
        df_ytd_matched["_key"] = df_ytd_matched[merge_key].apply(tuple, axis=1)
    
        # Drop duplicates to prevent cartesian matching 
        df_fy_matched = df_fy_matched.drop_duplicates()
        df_ytd_matched = df_ytd_matched.drop_duplicates()
    
        # Safe merge (1:1 expected now)
        df_merged = pd.merge(
            df_fy_matched,
            df_ytd_matched[["_key", "current_period_value", "prior_period_value"]],
            on="_key",
            suffixes=("_current", "_prior")
        )
    
        # === Calculate 4Q values ===
        df_merged["current_period_value"] = df_merged["current_period_value_current"] - df_merged["current_period_value_prior"]
        df_merged["prior_period_value"] = df_merged["prior_period_value_current"] - df_merged["prior_period_value_prior"]
    
        # === Construct df_4q_output with correct names ===
        df_4q_output = pd.DataFrame()
        df_4q_output["current_tag"] = df_merged["tag"]
        df_4q_output["current_date_type"] = df_merged["date_type"]
        df_4q_output["current_presentation_role"] = df_merged["presentation_role"]
    
        for col in AXIS_COLS:
            df_4q_output["current_" + col] = df_merged[col]
    
        df_4q_output["start_current"] = df_merged["start_current"]
        df_4q_output["end_current"] = pd.NaT
        df_4q_output["current_period_value"] = df_merged["current_period_value_current"]
        df_4q_output["contextref_current"] = None
    
        df_4q_output["prior_start"] = df_merged["start_current"]
        df_4q_output["prior_end"] = pd.NaT
        df_4q_output["prior_period_value"] = df_merged["prior_period_value_current"]
        df_4q_output["contextref_prior"] = None
    
        # === Standardize for downstream compatability
        df_4q_output = standardize_zip_output(df_4q_output)
    
        print(f"✅ Final 4Q output standardized: {len(df_4q_output)} rows")
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === IDENTIFY UNMATCHED ROWS AFTER EXACT MERGE ===
    
    if FOUR_Q_MODE:
    
        # Reuse the same merge key structure from exact match phase
        merge_key = ["tag"] + AXIS_COLS
    
        # FY rows that did NOT match in df_merged
        used_keys = set(df_merged["_key"])
        df_fy_unmatched = df_fy_matched[~df_fy_matched["_key"].isin(used_keys)]
    
        # YTD rows that did NOT match in df_merged
        df_ytd_unmatched = df_ytd_matched[~df_ytd_matched["_key"].isin(used_keys)]
    
        print(f"🔍 Unmatched FY rows: {len(df_fy_unmatched)}")
        print(f"🔍 Unmatched YTD rows: {len(df_ytd_unmatched)}")
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === FUZZY MATCH UNMATCHED ROWS ACROSS AXIS COLUMNS ===
    # Use fuzzy match as a fallback match
    
    from rapidfuzz import fuzz
    
    if FOUR_Q_MODE:
    
        fuzzy_matched_rows = []
    
        for i, row_fy in df_fy_unmatched.iterrows():
            for j, row_ytd in df_ytd_unmatched.iterrows():
    
                # === Match only if tag and start_current are exactly equal
                if row_fy["tag"] != row_ytd["tag"]:
                    continue
    
                # === Compare axis columns using partial_ratio
                passed = True
                for col in AXIS_COLS:
                    a = row_fy.get(col, "")
                    b = row_ytd.get(col, "")
    
                    if not (isinstance(a, str) and isinstance(b, str)):
                        passed = False
                        break
    
                    score = fuzz.partial_ratio(a, b)
                    if score < 80:
                        passed = False
                        break
    
                if passed:
                    # === Build fuzzy matched row (same format as df_merged)
                    fuzzy_matched_rows.append({
                        "tag": row_fy["tag"],
                        "date_type": row_fy["date_type"],
                        "start_current": row_fy["start_current"], #retained for reference, not used for merge
                        "end_current": pd.NaT,
                        "current_period_value_current": row_fy["current_period_value"],
                        "prior_period_value_current": row_fy["prior_period_value"],
                        "current_period_value_prior": row_ytd["current_period_value"],
                        "prior_period_value_prior": row_ytd["prior_period_value"],
                        "contextref_current": None,
                        "contextref_prior": None,
                        "presentation_role": row_fy["presentation_role"],
                        **{col: row_fy[col] for col in AXIS_COLS},
                        "_key": None  # optional: kept here as placeholder
                    })
                    break  # Stop after first match for each FY row
    
        # === Convert to DataFrame and append to df_merged
        df_fuzzy_merged = pd.DataFrame(fuzzy_matched_rows)
    
        print(f"✅ Added {len(df_fuzzy_merged)} fuzzy-matched rows to df_merged.")
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === COMBINE exact and fuzzy matches and finalize 4Q output
    
    if FOUR_Q_MODE:
        # Step 1: Combine exact + fuzzy
        df_merged = pd.concat([df_merged, df_fuzzy_merged], ignore_index=True)
    
        # Step 2: Calculate 4Q values
        df_merged["current_period_value"] = (
            df_merged["current_period_value_current"] - df_merged["current_period_value_prior"]
        )
        df_merged["prior_period_value"] = (
            df_merged["prior_period_value_current"] - df_merged["prior_period_value_prior"]
        )
    
        # Step 3: Standardize and clean for downstream use
        df_4q_output = standardize_zip_output(df_merged)
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === Audit Fuzzy Matches ===
    # Check the near-miss fuzzy matches (to make sure its correctly categorizing)
    
    from rapidfuzz import fuzz
    
    # === AUDIT: Fuzzy Near-Miss Logging ===
    if FOUR_Q_MODE:
    
        borderline_log = []
    
        for i, row_fy in df_fy_unmatched.iterrows():
            for j, row_ytd in df_ytd_unmatched.iterrows():
    
                if row_fy["tag"] != row_ytd["tag"]:
                    continue  # Only audit rows with same tag
    
                for col in AXIS_COLS:
                    a = row_fy.get(col, "")
                    b = row_ytd.get(col, "")
                    if not (isinstance(a, str) and isinstance(b, str)):
                        continue
    
                    score = fuzz.partial_ratio(a, b)
                    if 70 <= score < 80:
                        borderline_log.append({
                            "tag": row_fy["tag"],
                            "axis_column": col,
                            "FY_value": a,
                            "YTD_value": b,
                            "fuzzy_score": score
                        })
    
        # === Show audit result
        df_borderline_audit = pd.DataFrame(borderline_log)
        print(f"🔍 Borderline fuzzy matches (score 70–79): {len(df_borderline_audit)}")
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === Match Instants: Current Q vs Prior Q from 10-K ===
    
    if FOUR_Q_MODE:
        print("\n🏦 Matching instant facts (current_q vs prior_q) from 10-K...")
    
        # === Filter instants & current Q (full year in a 10-K) ===
        df_curr_inst = df_current_10k[df_current_10k["matched_category"] == "current_q"].copy()
        df_prior_inst = df_current_10k[df_current_10k["matched_category"] == "prior_q"].copy()
        
        # Filter for period_type = 'instant'
        df_curr_inst = df_curr_inst[df_curr_inst["period_type"] == "instant"].copy()
        df_prior_inst = df_prior_inst[df_prior_inst["period_type"] == "instant"].copy()
    
        # === Flatten presentation_role (some are lists)
        df_curr_inst["presentation_role"] = df_curr_inst["presentation_role"].apply(
            lambda x: "|".join(sorted(x)) if isinstance(x, list) else str(x)
        )
        df_prior_inst["presentation_role"] = df_prior_inst["presentation_role"].apply(
            lambda x: "|".join(sorted(x)) if isinstance(x, list) else str(x)
        )
    
        # === Fill axis values ===
        for col in AXIS_COLS:
            df_curr_inst[col] = df_curr_inst[col].fillna("__NONE__")
            df_prior_inst[col] = df_prior_inst[col].fillna("__NONE__")
    
        # === Match keys ===
        MATCH_COLS = ["tag", "presentation_role"] + AXIS_COLS
        MIN_KEYS = ["tag"]
    
        # === Trim
        df_curr_inst_trim = df_curr_inst[MATCH_COLS + ["value", "contextref"]].copy()
        df_prior_inst_trim = df_prior_inst[MATCH_COLS + ["value", "contextref"]].copy()
    
        # === Drop duplicates
        df_curr_inst_trim = df_curr_inst_trim.drop_duplicates()
        df_prior_inst_trim = df_prior_inst_trim.drop_duplicates()
    
        # === Adaptive match
        print("\n🧠 Matching current_q vs prior_q instants (balance sheet)...")
        match_keys = run_adaptive_match_keys(df_curr_inst_trim, df_prior_inst_trim, MATCH_COLS, MIN_KEYS)
    
        # === Zip match
        df_zip = zip_match_in_order(df_curr_inst_trim, df_prior_inst_trim, match_keys)
        df_instants = standardize_zip_output(df_zip)
    
        # === Preview & Log
        total_inst_curr = len(df_curr_inst_trim)
        matched_inst_rows = len(df_instants)
        match_rate_inst = matched_inst_rows / max(total_inst_curr, 1)
    
        print(f"✅ Instant match rate: {match_rate_inst * 100:.1f}%")
        log_metric("match_rate", {"instant": match_rate_inst})
    
    else:
        print("⚙️ Skipped: Not in 4Q mode.")
    
    
    # In[ ]:
    
    
    # === FINALIZE 4Q COMBINED OUTPUT ==============================
    
    if FOUR_Q_MODE:
        
        # Combine flow and instant value matches
        df_final_combined = pd.concat([
            df_4q_output, # standardized fuzzy+exact
            df_instants # from balance sheet
        ], ignore_index=True)
        
        # === CLEANUP STEPS ============================================
        
        # Drop rows where both values are missing
        df_final_combined = df_final_combined.dropna(
            subset=["current_period_value", "prior_period_value"],
            how="all"
        )
        # Drop exact duplicate rows (same values, same tag, etc.)
        df_final_combined = df_final_combined.drop_duplicates()
        
        # === PREVIEW & LOG RESULTS ============================================
    
        total_combined_rows = len(df_final_combined)
        total_full_year_facts = len(df_fy_curr) + len(df_curr_inst)
        match_rate_final_4q = total_combined_rows / total_full_year_facts
        print(f"✅ Final 4Q match rate: {match_rate_final_4q * 100:.1f}%")
    
        print(f"• Flow-based rows: {len(df_4q_output)}")
        print(f"• Instant rows: {len(df_instants)}")
        print(f"• Final 4Q combined output created: {len(df_final_combined)} rows")
    
    # Only log final_match_rate in pure 4Q mode (not FY mode)
    if FOUR_Q_MODE and not FULL_YEAR_MODE:
        log_metric("final_match_rate", match_rate_final_4q)
    
    
    # In[ ]:
    
    
    # === FULL YEAR WORKFLOW =============================================
    # === Match Full Year (current vs prior) from target 10-K ===
    
    if FOUR_Q_MODE and FULL_YEAR_MODE:
        
        print("\n📘 Matching full year facts (current_full_year vs prior_full_year) from 10-K...")
    
        # === Filter matched categories ===
        df_fy_curr = df_current_10k[df_current_10k["matched_category"] == "current_full_year"].copy()
        df_fy_prior = df_current_10k[df_current_10k["matched_category"] == "prior_full_year"].copy()
    
        # === Flatten presentation_role (some are lists)
        df_fy_curr["presentation_role"] = df_fy_curr["presentation_role"].apply(
            lambda x: "|".join(sorted(x)) if isinstance(x, list) else str(x)
        )
        df_fy_prior["presentation_role"] = df_fy_prior["presentation_role"].apply(
            lambda x: "|".join(sorted(x)) if isinstance(x, list) else str(x)
        )
    
        # === Fill axis values ===
        for col in AXIS_COLS:
            df_fy_curr[col] = df_fy_curr[col].fillna("__NONE__")
            df_fy_prior[col] = df_fy_prior[col].fillna("__NONE__")
    
        # === Match keys ===
        MATCH_COLS = ["tag", "presentation_role"] + AXIS_COLS
        MIN_KEYS = ["tag"]
    
        # === Trim
        df_fy_curr_trim = df_fy_curr[MATCH_COLS + ["start", "end", "value", "contextref", "date_type"]].copy()
        df_fy_prior_trim = df_fy_prior[MATCH_COLS + ["start", "end", "value", "contextref", "date_type"]].copy()
    
        # === Adaptive match
        match_keys = run_adaptive_match_keys(df_fy_curr_trim, df_fy_prior_trim, MATCH_COLS, MIN_KEYS)
    
        # === Zip match
        df_zip = zip_match_in_order(df_fy_curr_trim, df_fy_prior_trim, match_keys)
        df_full_year_matched = standardize_zip_output(df_zip)
    
        # === Preview results
        print(f"\n✅ Matched full-year rows (FY vs FY): {len(df_full_year_matched)}")
        
        total_fy_rows = len(df_fy_curr)
        matched_fy_rows = len(df_full_year_matched)
        match_rate_fy_final = matched_fy_rows / max(total_fy_rows, 1)
        
        log_metric("match_rate", {"fy_final": match_rate_fy_final})
        print(f"✅ FY final match rate: {match_rate_fy_final:.1%}")
    
    else:
        print("⚙️ Skipped: Not in 4Q mode.")
    
    
    # In[ ]:
    
    
    # === FULL-YEAR WORKFLOW =======================================
    # === Combine FY flow and instant facts for export ============
    
    if FOUR_Q_MODE and FULL_YEAR_MODE:
        print("\n📦 Building Full Year Matched Output...")
    
        # Combine full-year flow values + balance sheet instants
        df_final_fy = pd.concat([
            df_full_year_matched,
            df_instants
        ], ignore_index=True)
    
    # === CLEANUP STEPS ============================================
        
        # Drop rows where both values are missing (sanity)
        df_final_fy = df_final_fy.dropna(subset=["current_period_value", "prior_period_value"], how="all")
    
        # Remove exact duplicates (row-level)
        df_final_fy = df_final_fy.drop_duplicates()
    
        # Optional: deduplicate by value pair only
        df_final_fy = df_final_fy.drop_duplicates(subset=["current_period_value", "prior_period_value"])
    
    # === PREVIEW & LOG RATES ==================================================
    
        print(f"✅ Final FY Combined {len(df_final_fy)} rows")
    
        total_full_year_facts = len(df_fy_curr) + len(df_curr_inst)
        final_fy_row_count = len(df_final_fy)
        match_rate_fy_final_combined = final_fy_row_count / max(total_full_year_facts, 1)
    
        print(f"✅ Final full-year match rate: {match_rate_fy_final_combined:.1%}")
    
        if FOUR_Q_MODE and FULL_YEAR_MODE:
            log_metric("final_match_rate", match_rate_fy_final_combined)
            
    else:
        print("⚙️ Skipped: Not in full-year mode.")
    
    # TODO: log match diagnostics here (after modularization) - log match rate of different match steps
    
    
    # In[ ]:
    
    
    # === 4Q WORKFLOW =============================================
    # === Audit: Check what tags in the 10-K were missed in final export ===
    # Note: this is pre-deduplicated (number will skew higher)
    
    if FOUR_Q_MODE:
        print("\n📐 Auditing missing categorized tags by presentation role...")
    
        # Filter only categorized tags in df_current_10k
        df_categorized = df_current_10k[df_current_10k["date_type"].notna()].copy()
        tags_categorized = set(df_categorized["tag"])
        tags_in_final = set(df_final_combined["tag"])
    
        # Compute missing: categorized tags that didn’t make it into final
        missing_tags = tags_categorized - tags_in_final
        df_missing = df_categorized[df_categorized["tag"].isin(missing_tags)].copy()
    
        if not df_missing.empty:
            
            # Normalize missing presentation_role
            df_missing["presentation_role"] = df_missing["presentation_role"].fillna("(none)").str.strip().str.lower()
    
            # 3. Count missing by role
            summary = df_missing.groupby("presentation_role")["tag"].nunique().reset_index()
            summary = summary.rename(columns={"tag": "missing_tag_count"})
            summary = summary.sort_values(by="missing_tag_count", ascending=False)
    
            print(f"⚠️ Found {len(df_missing)} unmatched tags across {summary.shape[0]} roles.")
    
        else:
            print("✅ No missing tags. All current 10-K tags are present in final output.")
    else:
        print("⚙️ Skipped: Not in 4Q mode.")
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Structured match: Match Current Quarter vs Prior Quarter in the Current Target 10Q ===
    
    import pandas as pd
    import numpy as np
    
    # === CONFIG ===
    
    if not FOUR_Q_MODE:
        # === Normal 10-Q build
        
        MATCH_COLS = ["tag", "date_type"] + AXIS_COLS
    
        # Step 1: Filter just current_q and prior_q
        df_curr_q = df_current[df_current["matched_category"].isin(["current_q", "current_ytd"])].copy()
        df_prior_q = df_current[df_current["matched_category"].isin(["prior_q", "prior_ytd"])].copy()
    
        for col in AXIS_COLS:
            df_curr_q[col] = df_curr_q[col].fillna("__NONE__")
            df_prior_q[col] = df_prior_q[col].fillna("__NONE__")
        
        # Step 2: Trim to needed columns
        df_curr_trim = df_curr_q[MATCH_COLS + ["start", "end", "value", "contextref", "presentation_role"]].copy()
        df_prior_trim = df_prior_q[MATCH_COLS + ["start", "end", "value", "contextref", "presentation_role"]].copy()
    
        # === DIAGNOSTIC: Count duplicate match groups
        curr_group_sizes = df_curr_trim.groupby(MATCH_COLS).size()
        prior_group_sizes = df_prior_trim.groupby(MATCH_COLS).size()
        
        print("🔁 Duplicate match groups:")
        print(f"  • In current: {(curr_group_sizes > 1).sum()} groups with >1 row")
        print(f"  • In prior  : {(prior_group_sizes > 1).sum()} groups with >1 row")
    
        # === Optional: Match key overlap diagnostic
    
        curr_keys = set(df_curr_trim.groupby(MATCH_COLS).groups.keys())
        prior_keys = set(df_prior_trim.groupby(MATCH_COLS).groups.keys())
        shared_keys = curr_keys & prior_keys
    
        shared_ratio = len(shared_keys) / max(len(curr_keys), 1)
        print(f"\n🔍 Matching on: {MATCH_COLS}")
        print(f"   • Current keys: {len(curr_keys)}")
        print(f"   • Prior keys  : {len(prior_keys)}")
        print(f"   • Shared keys : {len(shared_keys)} ({shared_ratio:.2%} of current)")
    
        # Step 3: Zip match
        df_zip = zip_match_in_order(df_curr_trim, df_prior_trim, MATCH_COLS)
    
        # Step 3b: Rename and align to FINAL_COLS structure
        df_final = standardize_zip_output(df_zip)
        
        # Step 4: Clean up nulls and tag blanks
        df_final = df_final[df_final["tag"].notna()]
        df_final = df_final[df_final["tag"].str.strip() != ""]
    
        # Step 5: Preview and Log
        
        match_rate_q = len(df_final) / max(len(df_curr_q), 1)
        log_metric("match_rate", {"quarterly": match_rate_q})
        print(f"\n✅ Match rate (quarterly facts): {match_rate_q:.1%}")
        print(f"📊 Final total matched rows: {len(df_final)}")
    
    else:
        # 4Q Special build
        pass
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Structured match: YTD and Instant facts using the Prior Year's Quarterly Filing ===
    
    if not FOUR_Q_MODE:
        
        # === STEP 0: Set column names and match criteria
    
        MATCH_COLS_YTD = ["tag", "date_type"] + AXIS_COLS
        MATCH_COLS_INSTANT = ["tag", "end", "date_type", "presentation_role"] + AXIS_COLS
    
        # === Minimum fallback keys
        MIN_MATCH_COLS_YTD = ["tag", "date_type"]
        MIN_MATCH_COLS_INSTANT = ["tag", "end", "date_type"]
    
        # === STEP 0: Standardize fill values
        
        for col in AXIS_COLS:
            df_current[col] = df_current[col].fillna("__NONE__")
            df_prior[col] = df_prior[col].fillna("__NONE__")
        
        # === STEP 0: Filter df_current for current_q and current_YTD
        
        df_curr_filtered = df_current[df_current["matched_category"].isin(["current_q", "current_ytd"])].copy()
        print(f"✅ Filtered df_current to {len(df_curr_filtered)} rows (current_q, current_ytd)")
    
        # Split into YTD and Instant for separate merges
        
        df_ytd = df_curr_filtered[df_curr_filtered["date_type"] == "YTD"].copy()
        df_instant = df_curr_filtered[df_curr_filtered["period_type"] == "instant"].copy()
        print(f"YTD rows: {len(df_ytd)}, Instant rows: {len(df_instant)}")
        
        # === STEP 2: Match YTD facts using prior y/y quarterly filing
    
        df_ytd_curr = df_ytd[df_ytd["matched_category"] == "current_ytd"].copy()
        df_ytd_prior = df_prior[df_prior["matched_category"] == "current_ytd"].copy()  # Prior Q stores prior YTD as "current_ytd"
        
        # Trim to match keys + data
        df_ytd_curr_trim = df_ytd_curr[MATCH_COLS_YTD + ["start", "end", "value", "contextref", "presentation_role"]].copy()
        df_ytd_prior_trim = df_ytd_prior[MATCH_COLS_YTD + ["start", "end", "value", "contextref", "presentation_role"]].copy()
        
        # === DIAGNOSTIC: Count duplicate match groups
        curr_ytd_group_sizes = df_ytd_curr_trim.groupby(MATCH_COLS_YTD).size()
        prior_ytd_group_sizes = df_ytd_prior_trim.groupby(MATCH_COLS_YTD).size()
        
        print(f"🔁 Duplicate YTD match groups:")
        print(f"  • In current: {(curr_ytd_group_sizes > 1).sum()} groups with >1 row")
        print(f"  • In prior  : {(prior_ytd_group_sizes > 1).sum()} groups with >1 row")
    
        df_ytd_curr_trim = df_ytd_curr_trim.drop_duplicates()
        df_ytd_prior_trim = df_ytd_prior_trim.drop_duplicates()
    
        print("\n🧠 YTD: Starting adaptive match key evaluation...")
        
        match_cols_ytd = run_adaptive_match_keys(df_ytd_curr_trim, df_ytd_prior_trim, MATCH_COLS_YTD, MIN_MATCH_COLS_YTD)
        df_ytd_merged = zip_match_in_order(df_ytd_curr_trim, df_ytd_prior_trim, match_cols_ytd)
        df_ytd_merged = standardize_zip_output(df_ytd_merged)
    
        # === STEP 3: Match Instant facts from prior y/y quarterly filing
    
        df_instant_curr = df_instant[df_instant["matched_category"] == "current_q"].copy()
        df_instant_prior = df_prior[
            (df_prior["matched_category"] == "current_q") &
            (df_prior["period_type"] == "instant")
        ].copy()
    
        #Date shifting logic
        
        # Step 1: Calculate the true delta between fiscal year-ends (in days)
        fye_curr = parse_date(target_10q["document_period_end"])
        fye_prior = parse_date(prior_10q["document_period_end"])
        year_delta = (fye_curr - fye_prior).days
        
        # Step 2: Add that exact day offset to prior instant 'end' values
        df_instant_prior["end"] = df_instant_prior["end"].apply(lambda x: x + pd.Timedelta(days=year_delta))
    
        #Turn presentation role data into string
        df_instant_curr["presentation_role"] = df_instant_curr["presentation_role"].apply(
            lambda x: "|".join(sorted(x)) if isinstance(x, list) else str(x)
        )
        df_instant_prior["presentation_role"] = df_instant_prior["presentation_role"].apply(
            lambda x: "|".join(sorted(x)) if isinstance(x, list) else str(x)
        )
    
        #Create two match groups for sequential match
        df_instant_curr_trim = df_instant_curr[MATCH_COLS_INSTANT + ["value", "contextref"]].copy()
        df_instant_prior_trim = df_instant_prior[MATCH_COLS_INSTANT + ["value", "contextref"]].copy()
    
        # === DIAGNOSTIC: Count how many match groups have duplicate rows
        
        curr_group_sizes = df_instant_curr_trim.groupby(MATCH_COLS_INSTANT).size()
        prior_group_sizes = df_instant_prior_trim.groupby(MATCH_COLS_INSTANT).size()
        
        curr_dup_groups = (curr_group_sizes > 1).sum()
        prior_dup_groups = (prior_group_sizes > 1).sum()
        
        print(f"🔁 Duplicate match groups:")
        print(f"  • In current: {curr_dup_groups} groups with >1 row")
        print(f"  • In prior  : {prior_dup_groups} groups with >1 row")
    
        df_instant_curr_trim = df_instant_curr_trim.drop_duplicates()
        df_instant_prior_trim = df_instant_prior_trim.drop_duplicates()
    
        print("\n🧠 Instant: Starting adaptive match key evaluation...")
    
        match_cols_instant = run_adaptive_match_keys(df_instant_curr_trim, df_instant_prior_trim, MATCH_COLS_INSTANT, MIN_MATCH_COLS_INSTANT)
        df_instant_merged = zip_match_in_order(df_instant_curr_trim, df_instant_prior_trim, match_cols_instant)
        df_instant_merged = standardize_zip_output(df_instant_merged)
        
        print(f"✅ Instant facts matched (zip): {len(df_instant_merged)}")
        
        # === STEP 4: Append to df_final
        df_final = pd.concat(
            [df_final] + [f for f in [df_ytd_merged, df_instant_merged] if not f.empty],
            ignore_index=True
        )
        
        # === FINAL CLEANUP
        df_final = df_final.dropna(subset=["current_period_value", "prior_period_value"], how="any")
        df_final = df_final.drop_duplicates()
        print(f"✅ Final shape after dropping exact duplicates: {df_final.shape}")
    
        # === YTD Match Rate ===
        match_rate_ytd = len(df_ytd_merged) / max(len(df_ytd_curr_trim), 1)
    
        # === Instant Match Rate ===
        match_rate_instants = len(df_instant_merged) / max(len(df_instant_curr_trim), 1)
    
        # === Log both match rates
        log_metric("match_rate", {
            "ytd": match_rate_ytd,
            "instants": match_rate_instants
        })
        
        # === Optional: Preview ===
        print("\n✅ Filled missing prior_period_value from prior filing.")
        
        print(f"➕ YTD matches added: {len(df_ytd_merged)}")
        print(f"➕ Instant matches added: {len(df_instant_merged)}")
    
        print(f"✅ YTD match rate: {match_rate_ytd:.1%}")
        print(f"✅ Instant match rate: {match_rate_instants:.1%}")
        
        print("\n✅ Appended YTD and Instant fact matches to df_final")
        print(f"📊 Final total matched rows: {len(df_final)}")
    
        num_duplicates = df_final.duplicated(keep=False).sum()
        print(f"🔁 Total exact duplicate rows in df_final: {num_duplicates}")
    
    else:
        # === 4Q Special build
        pass
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Deeper Audit: Check if any tags in the current 10-Q were missed  ===
    
    if not FOUR_Q_MODE:
    
        # 1. All tags from df_current
        all_tags_in_current = set(df_current["tag"])
        
        # 2. Tags in the final export table
        tags_in_final = set(df_final["tag"])
        
        # 3. Find tags that exist in current but NOT in final
        missing_tags = all_tags_in_current - tags_in_final
        
        # === Output ===
        if missing_tags:
            print(f"\n⚡ {len(missing_tags)} tags exist in current 10-Q but were missing from df_final:")
            for tag in sorted(missing_tags):
                print(f"  - {tag}")
    
            # ✅ Build DataFrame before saving
            df_missing_tags = df_current[df_current["tag"].isin(missing_tags)].copy()
    
            # Keep relevant context columns
            df_missing_tags = df_missing_tags[[
                "tag", "value", "start", "end",
            ]].drop_duplicates()
    
            df_missing_tags["status_note"] = ""  # optional column for manual review            
    
        else:
            print("\n✅ No missing tags. All current_q tags are properly included in df_final.")
    
    else:
        # === 4Q Special build
        pass
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Select facts in df_current where tag is in missing tags to check values
    
    if not FOUR_Q_MODE:
    
        df_missing_facts = df_current[df_current["tag"].isin(missing_tags)]
        
        # === Optional: Show a quick preview
        print(f"\n✅ Found {len(df_missing_facts)} facts for missing tags.")
    
    else:
        # === 4Q Special build
        pass
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Fallback Match: Unmatched tags (YTD + Instant) with Current / Prior 10-Q ====================================
    
    if not FOUR_Q_MODE:
    
        fallback_keys = ["tag", "end", "date_type"] + AXIS_COLS  # Looser match (no presentation_role)
    
        # Filter to fallback tags
        df_curr_fallback = df_current[df_current["tag"].isin(missing_tags)].copy()
        df_prior_fallback = df_prior[df_prior["tag"].isin(missing_tags)].copy()
    
        # Filter to keep only matched_category current_q or current_ytd
        df_curr_fallback = df_curr_fallback[
            (df_curr_fallback["matched_category"] == "current_q") |
            (df_curr_fallback["matched_category"] == "current_ytd")
        ]
        df_prior_fallback = df_prior_fallback[
            (df_prior_fallback["matched_category"] == "current_q") |
            (df_prior_fallback["matched_category"] == "current_ytd")
        ]
    
        # Drop any entries without keys
        df_curr_fallback = df_curr_fallback.dropna(subset=fallback_keys)
        df_prior_fallback = df_prior_fallback.dropna(subset=fallback_keys)
    
        # Fill axis cols
        for col in AXIS_COLS: 
            df_curr_fallback[col] = df_curr_fallback[col].fillna("__NONE__")
            df_prior_fallback[col] = df_prior_fallback[col].fillna("__NONE__")
    
        # Shift dates ONLY for instant prior facts
        
        # === Step 1: Calculate the true day offset between fiscal year ends
        fye_curr = parse_date(target_10q["fiscal_year_end"])
        fye_prior = parse_date(prior_10q["fiscal_year_end"])
        year_delta = (fye_curr - fye_prior).days
        
        # === Step 2: Shift only instant-type rows in prior fallback
        mask = df_prior_fallback["period_type"] == "instant"
        df_prior_fallback.loc[mask, "end"] = df_prior_fallback.loc[mask, "end"] + pd.Timedelta(days=year_delta)
    
        # Run fallback match
        df_fallback_matches = zip_match_in_order(df_curr_fallback, df_prior_fallback, fallback_keys)
    
        print(f"\n✅ Fallback matches generated: {len(df_fallback_matches)} rows")
        
        # Clean final output
        df_fallback_clean = standardize_zip_output(df_fallback_matches)
        df_fallback_clean = df_fallback_clean.dropna(subset=["current_period_value", "prior_period_value"], how="all")
        df_fallback_clean = df_fallback_clean.drop_duplicates()
        
        # Log fallback match rate
        fallback_match_rate = len(df_fallback_clean) / max(len(df_curr_fallback), 1)
        log_metric("match_rate", {"fallback": fallback_match_rate})
        print(f"✅ Fallback match rate: {fallback_match_rate:.1%}")
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Collision Audit for Fallback Matches ===
    
    if not FOUR_Q_MODE:
    
        flagged_fallback_df = audit_value_collisions(df_fallback_clean)
        flagged_fallback_values = set(flagged_fallback_df["current_period_value"])
    
        df_fallback_clean["collision_flag"] = df_fallback_clean["current_period_value"].apply( 
            lambda x: 1 if x in flagged_fallback_values else 0
        )
    
        if not flagged_fallback_df.empty:
            print(f"⚠️ Fallback collision detected — {len(flagged_fallback_df)} flagged rows")
    
        else:
            print("✅ No collision flags in fallback output")
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Audit to check for duplicate prior year values in fallback and final output ====================================
    
    if not FOUR_Q_MODE:
        
        # Look for cases where the same prior value is matched to multiple current values
        
        # Ensure relevant columns are present and drop rows with missing values
        df_final_check = df_final[["tag", "current_period_value", "prior_period_value"]].dropna(subset=["prior_period_value"]).copy()
        df_fallback_check = df_fallback_clean[["tag", "current_period_value", "prior_period_value"]].dropna(subset=["prior_period_value"]).copy()
        
        # Convert prior values to numeric for safe matching
        df_final_check["prior_period_value"] = pd.to_numeric(df_final_check["prior_period_value"], errors="coerce")
        df_fallback_check["prior_period_value"] = pd.to_numeric(df_fallback_check["prior_period_value"], errors="coerce")
        
        # Identify overlaps by prior_period_value
        overlap = pd.merge(
            df_final_check,
            df_fallback_check,
            on="prior_period_value",
            suffixes=("_final", "_fallback")
        )
            
        if not overlap.empty:
            print(f"🔍 Found {len(overlap)} overlapping prior values between main and fallback outputs.")
    
        else:
            print("✅ No overlapping prior values found.")
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Audit to check for value mismatch between matched fallback and final outputs ====================================
    
    if not FOUR_Q_MODE:
        # Check for shared prior period values that were matched to different current period values in final vs. fallback datasets
        
        # Ensure aligned and numeric
        df_final_check = df_final[["tag", "current_period_value", "prior_period_value"]].dropna(subset=["prior_period_value"]).copy()
        df_fallback_check = df_fallback_clean[["tag", "current_period_value", "prior_period_value"]].dropna(subset=["prior_period_value"]).copy()
        
        df_final_check["prior_period_value"] = pd.to_numeric(df_final_check["prior_period_value"], errors="coerce")
        df_fallback_check["prior_period_value"] = pd.to_numeric(df_fallback_check["prior_period_value"], errors="coerce")
        df_final_check["current_period_value"] = pd.to_numeric(df_final_check["current_period_value"], errors="coerce")
        df_fallback_check["current_period_value"] = pd.to_numeric(df_fallback_check["current_period_value"], errors="coerce")
        
        # Merge to find overlapping prior values
        overlap = pd.merge(
            df_final_check,
            df_fallback_check,
            on="prior_period_value",
            suffixes=("_final", "_fallback")
        )
        
        # Find rows where current_period_value differs
        mismatches = overlap[
            overlap["current_period_value_final"] != overlap["current_period_value_fallback"]
        ]
        
        print(f"🔍 Found {len(mismatches)} mismatched current values for overlapping prior values.")
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Finalize fallback matches by removing overlapping prior values ============================
    
    if not FOUR_Q_MODE:
            
        # Step 1: Identify overlapping prior_period_values from earlier audit
        overlap_prior_values = set(overlap["prior_period_value"])
        
        # Step 2: Filter fallback matches to only keep non-duplicates
        df_fallback_unique = df_fallback_clean[
            ~df_fallback_clean["prior_period_value"].isin(overlap_prior_values)
        ].copy()
        
        print(f"✅ Result: {len(df_fallback_unique)} fallback matches added after removing {len(overlap_prior_values)} overlapping prior values.")
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Append fallback matches to df_final before visual export ================
    
    if not FOUR_Q_MODE:
        
        # Ensure it's aligned to final structure
        df_fallback_finalized = df_fallback_unique.copy()
        df_fallback_finalized = standardize_zip_output(df_fallback_finalized)
    
        if not df_fallback_finalized.empty:
            # Append fallback to final
            df_final = pd.concat([df_final, df_fallback_finalized], ignore_index=True)
            df_final.reset_index(drop=True, inplace=True)
            
            print(f"✅ Final row count after fallback merge: {len(df_final)}")
    
        else:
            print("⚠️ df_fallback_finalized is empty — skipped appending to df_final.")
    
        # === Final match rate logging
        total_facts_current = len(df_current[df_current["matched_category"].isin(["current_q", "current_ytd"])])
        match_rate_final_quarter = len(df_final) / max(total_facts_current, 1)
    
        print(f"✅ Final quarterly match rate: {match_rate_final_quarter:.1%}")
        log_metric("final_match_rate", match_rate_final_quarter)
    
    # TODO: log match diagnostics here (after modularization) - log match rate of different match steps
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Audit: New disclosures (tag + axis) not seen in prior year's Q ===
    
    if not FOUR_Q_MODE:
        print("\n📌 Auditing NEW disclosures in current quarter (tag + axis)...")
    
        MATCH_KEYS = ["tag"] + AXIS_COLS  # e.g., tag + segment, geo, product, etc.
    
        # Step 1: Build disclosure keys for current and prior
        curr_keys = set(df_final[MATCH_KEYS].apply(tuple, axis=1))
        prior_keys = set(df_prior[MATCH_KEYS].apply(tuple, axis=1))  # from prior 10-Q enrichment
    
        # Step 2: Identify new keys
        new_keys = curr_keys - prior_keys
    
        # Step 3: Filter rows where match_key is new
        df_final["_match_key"] = df_final[MATCH_KEYS].apply(tuple, axis=1)
        df_new_disclosures = df_final[df_final["_match_key"].isin(new_keys)].drop(columns=["_match_key"])
    
        # Step 4: Preview
        print(f"🆕 Found {len(df_new_disclosures)} new disclosures this quarter.")
        print("\n🔍 Sample of new disclosure tags:\n")
        print(df_new_disclosures["tag"].dropna().unique()[:20])  # Adjust slice as needed
    
    else:
        print("⚙️ Skipped: Not in quarterly mode.")
    
    
    # In[ ]:
    
    
    # === NORMAL 10-Q WORKFLOW ====================================
    # === Identify Facts (currentQ and YTD) That Did Not Make Final Output ===
    
    if not FOUR_Q_MODE:
        
        # Step 1: Filter df_current to current_q and current_ytd with non-null values
        df_qytd_current = df_current[
            df_current["matched_category"].isin(["current_q", "current_ytd"]) &
            df_current["value"].notna()
        ].copy()
        
        # Step 2: Build minimal comparison key: tag + value + end date with temporary column
        df_qytd_current["_val_key"] = df_qytd_current[["tag", "value", "end"]].astype(str).agg("|".join, axis=1)
        df_final["_val_key"] = df_final[["tag", "current_period_value", "end_current"]].astype(str).agg("|".join, axis=1)
        
        # Step 3: Identify rows in df_current that are NOT in df_final
        unmatched_keys = ~df_qytd_current["_val_key"].isin(df_final["_val_key"])
        df_new_disclosures = df_qytd_current[unmatched_keys].drop(columns=["_val_key"])
    
        # Standardize presentation_role
        df_new_disclosures["presentation_role"] = df_new_disclosures["presentation_role"].fillna("(none)").str.strip().str.lower()
    
        # Step 4: Summary by presentation role
        summary = df_new_disclosures["presentation_role"].value_counts().reset_index()
        summary.columns = ["presentation_role", "new_disclosure_count"]
    
        # Output results
        # === Summary of unmatched current_q/ytd disclosures ===
        total_qytd_facts = len(df_qytd_current)
        unmatched_facts = len(df_new_disclosures)
        unmatched_pct = unmatched_facts / max(total_qytd_facts, 1)
        
        print(f"✅ New unmatched current_q/ytd disclosures: {unmatched_facts} out of {total_qytd_facts} ({unmatched_pct:.1%})")
    
    
    # In[ ]:
    
    
    # === SHARED: Collision Audit  ===
    # === Flag duplicate values that have multiple matches ===
    
    # Select appropriate final output
    if FOUR_Q_MODE and not FULL_YEAR_MODE:
        audit_df = df_final_combined
    elif FOUR_Q_MODE and FULL_YEAR_MODE:
        audit_df = df_final_fy
    else:
        audit_df = df_final
    
    # Run audit
    flagged_df = audit_value_collisions(audit_df)
    flagged_values = set(flagged_df["current_period_value"])
    audit_df["collision_flag"] = audit_df["current_period_value"].apply(lambda x: 1 if x in flagged_values else 0)
    
    # === Diagnostics Summary ===
    num_total_rows = len(audit_df)
    num_unique_collisions = flagged_df[["tag", "current_period_value", "prior_period_value"]].drop_duplicates().shape[0]
    collision_rate = num_unique_collisions / max(num_total_rows, 1)
    
    print(f"🔁 Unique (tag, current, prior) collision combinations: {num_unique_collisions}")
    print(f"📊 Collision rate: {collision_rate:.1%}")
    log_metric("collision_rate", collision_rate)
    
    
    # In[ ]:
    
    
    # === SHARED LOGIC (Apply Visual Logic and Export Dataframe) =============
    # === Create New DataFrame with Visual Values Using Company Negate Labels ===
    
    # === Flip signs for negated tags in presentation filing
    
    def apply_visual_signs(df, negated_tags):
        """
        Applies visual sign-flipping to financial values for tags that are commonly reported as negative 
        (e.g., expenses, losses) but are conceptually positive.
    
        Args:
            df (DataFrame): DataFrame containing "tag", "current_period_value", and "prior_period_value".
            negated_tags (set or list): Collection of XBRL tags that should be visually flipped 
                                        (e.g., ["us-gaap:OperatingExpenses", "us-gaap:InterestExpense"]).
    
        Returns:
            DataFrame: Same DataFrame with two new columns:
                - 'visual_current_value': current value with visual sign flip applied if tag in negated_tags
                - 'visual_prior_value': prior value with visual sign flip applied if tag in negated_tags
    
        Example:
            apply_visual_signs(df, {"us-gaap:OperatingExpenses"})
        """
        
        df["visual_current_value"] = df.apply(
            lambda row: -row["current_period_value"] if row["tag"] in negated_tags and pd.notna(row["current_period_value"]) else row["current_period_value"],
            axis=1
        )
        df["visual_prior_value"] = df.apply(
            lambda row: -row["prior_period_value"] if row["tag"] in negated_tags and pd.notna(row["prior_period_value"]) else row["prior_period_value"],
            axis=1
        )
        return df
    
    # === Apply visual logic + export based on mode
    
    if FOUR_Q_MODE and FULL_YEAR_MODE:
        print("📦 FY mode detected — applying negated signs and exporting CSV.")
        export_df = apply_visual_signs(df_final_fy.copy(), negated_tags)
        filename = f"{TICKER}_{annual_label}_final_visual.csv"
    
    elif FOUR_Q_MODE:
        print("📦 4Q mode detected — applying negated signs and exporting CSV.")
        export_df = apply_visual_signs(df_final_combined.copy(), negated_tags)
        filename = f"{TICKER}_{target_label}_final_visual.csv"
    
    else:
        print("📦 10-Q mode detected — applying negated signs and exporting CSV.")
        export_df = apply_visual_signs(df_final.copy(), negated_tags)
        filename = f"{TICKER}_{target_label}_final_visual.csv"
    
    # === Count how many tags had their sign flipped
    
    flipped_current = export_df.apply(
        lambda row: row["tag"] in negated_tags and pd.notna(row["current_period_value"]) and row["visual_current_value"] != row["current_period_value"],
        axis=1
    ).sum()
    
    flipped_prior = export_df.apply(
        lambda row: row["tag"] in negated_tags and pd.notna(row["prior_period_value"]) and row["visual_prior_value"] != row["prior_period_value"],
        axis=1
    ).sum()
    
    log_metric("sign_flip_count", {
        "current": int(flipped_current),
        "prior": int(flipped_prior)
    })
    
    print(f"🔁 Sign flip applied on:")
    print(f"   • Current period: {flipped_current} values")
    print(f"   • Prior period  : {flipped_prior} values")
    
    # === Sort export_df by presentation_role and tag in-place
    export_df = export_df.sort_values(by=["presentation_role", "tag"])
    
    # Remove exact duplicate rows across current/prior values
    export_df = export_df.drop_duplicates(subset=["current_period_value", "prior_period_value"])
    
    
    # In[ ]:
    
    
    # === FINAL EXPORTS TO MODEL ==================================
    # === Export values to updater file
    
    # === Set Export Folder and Filename
    
    os.makedirs(EXPORT_UPDATER_DIR, exist_ok=True)
    export_updater_filename = f"{TICKER}_{QUARTER}Q{str(YEAR)[-2:]}_{EXCEL_FILE}"
    export_updater_path = os.path.join(EXPORT_UPDATER_DIR, export_updater_filename)
    
    # === Load the Updater workbook
    updater_path = EXCEL_FILE  # Adjust if different
    wb = openpyxl.load_workbook(updater_path, keep_vba=True)
    
    # === Select the Raw_data sheet
    sheet = wb["Raw_data"]
    
    # === Clear old data (optional, but safe to do)
    for row in sheet["A2:E5000"]:
        for cell in row:
            cell.value = None
    
    # === Write metadata in columns G–H (optional)
    sheet["G1"] = "Ticker"
    sheet["H1"] = TICKER
    sheet["G2"] = "Year"
    sheet["H2"] = YEAR
    sheet["G3"] = "Quarter"
    sheet["H3"] = QUARTER
    sheet["G4"] = "Full Year Mode"
    sheet["H4"] = str(FULL_YEAR_MODE)
    
    # === Choose the correct dataframe
    if FOUR_Q_MODE and FULL_YEAR_MODE:
        print("\n📦 FY mode — exporting full FY combined table.")
        export_df = df_final_fy
    elif FOUR_Q_MODE:
        print("\n📦 4Q mode — exporting full 4Q combined table.")
        export_df = df_final_combined
        
    else:
        print("\n📦 Normal 10-Q mode — exporting full quarterly combined table.")
        #export_df = df_final
    
    # === STRONG final clean before export ===
    
    # Drop any row where all columns are NaN (including tag)
    export_df = export_df.dropna(how="all")
    
    # Also drop rows where tag is blank string
    export_df = export_df[export_df["tag"].notna()]
    export_df = export_df[export_df["tag"].str.strip() != ""]
    
    # Also (optional) drop rows where both current/prior are missing
    export_df = export_df.dropna(subset=["current_period_value", "prior_period_value"], how="all")
    
    # Remove exact duplicate rows across current/prior values
    export_df = export_df.drop_duplicates(subset=["current_period_value", "prior_period_value"])
    
    # === Reset index so idx=0,1,2,3... and sort by presentation role
    export_df = export_df.reset_index(drop=True)
    
    # === Paste the selected DataFrame into Updater
    for idx, row in export_df.iterrows():
        sheet.cell(row=idx+2, column=1, value=row["tag"])
        sheet.cell(row=idx+2, column=2, value=row.get("visual_current_value", row.get("current_period_value")))
        sheet.cell(row=idx+2, column=3, value=row.get("visual_prior_value", row.get("prior_period_value")))
        sheet.cell(row=idx+2, column=4, value=row.get("presentation_role", ""))  # Presentation information
        sheet.cell(row=idx+2, column=5, value=row.get("collision_flag", 0))      # Collision flag
        
    # === Save the workbook to export folder
    wb.save(export_updater_path)
    print(f"📁 Updater file saved to: {export_updater_path}")
    
    wb.save(updater_path)
    print(f"📁 Updater file also saved to: {updater_path}")
    
    
    # === Export summary
    
    print(f"\n📄 Export summary: {QUARTER}Q {YEAR} data from {TICKER} ({CIK}) successfully written to {EXCEL_FILE}")
    print(f"✅ Data written to sheet Raw_data starting from A2.")
    print(f"📊 Total rows: {len(export_df)}")
    
    # === Show filing references for quarterly mode
    
    if not FOUR_Q_MODE and target_10q:
    
        print("\n🔗 Target 10-Q Filing Summary:")
        print(f"  • Ticker     : {TICKER}")
        print(f"  • CIK        : {CIK}")
        print(f"  • Form       : {target_10q.get('form', 'N/A')}")
        print(f"  • Label      : {target_10q.get('label', 'N/A')}")
        print(f"  • Period End : {target_10q.get('document_period_end', 'N/A')}")
        print(f"  • Accession  : {target_10q.get('accession', 'N/A')}")
        print(f"  • URL        : {target_10q.get('url', 'N/A')}")
        
        if DEBUG_MODE:
            print("\n🔗 Target 10-Q Filing Reference:")
            print(f"✅ {target_10q['label']} | Period End: {target_10q['document_period_end']} | URL: {target_10q['url']}")
            
            if prior_10q:
                print(f"\n✅ Found prior 10-Q:")
                print(f"✅ {prior_10q['label']} | Period End: {prior_10q['document_period_end']} | URL: {prior_10q['url']}")
    
    # === Show filing references for 4Q mode
    
    if FOUR_Q_MODE and not FULL_YEAR_MODE:
    
        print("\n🔗 Target 10-K Filing Summary:")
        print(f"  • Ticker     : {TICKER}")
        print(f"  • CIK        : {CIK}")
        print(f"  • Form       : {target_10k['form']}")
        print(f"  • Period End : {target_10k['document_period_end']}")
        print(f"  • Accession  : {target_10k['accession']}")
        print(f"  • URL        : {target_10k['url']}")
    
        if DEBUG_MODE:
    
            print("\n🔗 Target 10-K Filing Reference:")
            print(f"✅ Period End: {target_10k['document_period_end']} | URL: {target_10k['url']}")
            
            print("\n🔗 Current Year 10-Qs (Q1–Q3):")
            if q1_entry:
                print(f"  - Q1: {q1_entry['document_period_end']} | URL: {q1_entry['url']}")
        
            if q2_entry:    
                print(f"  - Q2: {q2_entry['document_period_end']} | URL: {q2_entry['url']}")
                
            if q3_entry:   
                print(f"  - Q3: {q3_entry['document_period_end']} | URL: {q3_entry['url']}")
        
            print("\n🔗 Prior Year 10-K Filing Reference:")
            print(f"✅ Period End: {prior_10k['document_period_end']} | URL: {prior_10k['url']}")
        
            print("\n🔗 Prior Year 10-Qs (Q1–Q3):")
            if q1_prior_entry:
                print(f"  - Q1: {q1_prior_entry['document_period_end']} | URL: {q1_prior_entry['url']}")
            
            if q2_prior_entry:
                print(f"  - Q2: {q2_prior_entry['document_period_end']} | URL: {q2_prior_entry['url']}")
            
            if q3_prior_entry:
                print(f"  - Q3: {q3_prior_entry['document_period_end']} | URL: {q3_prior_entry['url']}")
    
    # === FY Reference Summary for 4Q and Full Year ===
        
    if FOUR_Q_MODE and FULL_YEAR_MODE:
    
        print("\n🔗 Target 10-K Filing Summary:")
        print(f"  • Ticker     : {TICKER}")
        print(f"  • CIK        : {CIK}")
        print(f"  • Form       : {target_10k['form']}")
        print(f"  • Period End : {target_10k['document_period_end']}")
        print(f"  • Accession  : {target_10k['accession']}")
        print(f"  • URL        : {target_10k['url']}")
    
        if DEBUG_MODE:
            print("\n🔗 Target 10-K Filing Reference:")
            print(f"✅ Period End: {target_10k['document_period_end']} | URL: {target_10k['url']}")
    
    
    # In[ ]:
    
    
    # === SHARED LOGIC (e.g. negated labels, exports) =============
    #Check dataframe export is clean and complete
    
    if DEBUG_MODE:
        print(f"🔍 Exported DataFrame contains {export_df.shape[0]} rows × {export_df.shape[1]} columns.")
        print(f"🔍 Rows with missing tag: {(export_df['tag'].isna() | (export_df['tag'].str.strip() == '')).sum()}")
        print(f"🔍 Rows where both current and prior are missing: {(export_df['current_period_value'].isna() & export_df['prior_period_value'].isna()).sum()}")
    
    
    # In[ ]:
    
    
    from datetime import datetime
    import json
    
    # === Add end timestamp to metrics ===
    metrics["end_time"] = datetime.now().isoformat()
    
    # === Calculating processing time ===
    start = datetime.fromisoformat(metrics["start_time"])
    end = datetime.fromisoformat(metrics["end_time"])
    duration = (end - start).total_seconds()
    print(f"⏱️ Total processing time: {duration:.2f} seconds")
    log_metric("total_processing_seconds", duration)
    
    # === Export summary to JSON ===
    os.makedirs(OUTPUT_METRICS_DIR, exist_ok=True)
    summary_path = os.path.join(OUTPUT_METRICS_DIR, f"{TICKER}_{QUARTER}Q{str(YEAR)[-2:]}_summary_metrics.json")
    
    with open(summary_path, "w") as f:
        json.dump(metrics, f, indent=2)
    
    print(f"✅ Exported summary metrics to: {summary_path}")
    
    
    # In[ ]:
    
    
    print(f"\n📊 Final Metrics Dictionary:\n{json.dumps(metrics, indent=2)}")
    
    
    # In[ ]:
    
    
    
    
    
    # In[ ]:














    
    

    
    
    

    








    
    
    

    
    


